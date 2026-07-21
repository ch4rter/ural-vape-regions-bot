import re
import shutil
import sqlite3
import tempfile
from copy import copy
from collections import defaultdict
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


WAREHOUSES = {
    "center": "Москва",
    "west": "Санкт-Петербург",
    "ural": "Челябинск",
}


def normalize_price_text(value: str) -> str:
    value = value.casefold().replace("ё", "е")
    value = re.sub(r"[^a-zа-я0-9]+", " ", value)
    aliases = {
        "огго": "oggo",
        "влик": "vliq",
        "доджо": "dojo",
        "аромамикс": "ароматизатор",
        "аромамиксы": "ароматизатор",
        "арома": "ароматизатор",
        "жижка": "жидкость",
    }
    return " ".join(aliases.get(word, word) for word in value.split())


def clean_group_name(full_path: str) -> str:
    name = full_path.rsplit("/", 1)[-1].strip()
    name = re.sub(r"^\d+\s*\.\s*", "", name)
    name = re.sub(r"^Д\s+(?=[A-ZА-ЯЁ])", "", name)
    return re.sub(r"\s+", " ", name).strip()


def variant_display_name(name: str) -> str:
    name = re.sub(r"^\s*АКЦИЯ\s+", "", name, flags=re.IGNORECASE).strip()
    if " - " in name:
        return name.split(" - ", 1)[1].strip()
    flavor_match = re.search(r"\bс ароматом\s+(.+)$", name, flags=re.IGNORECASE)
    if flavor_match:
        return flavor_match.group(1).strip()
    return name


def category_from_path(full_path: str) -> tuple[str, str]:
    normalized = normalize_price_text(full_path)
    rules = (
        ("жидкост", "liquids", "Жидкости"),
        ("конструктор", "mixes", "Конструкторы и ароматизаторы"),
        ("однораз", "disposables", "Одноразовые системы"),
        ("картридж", "cartridges", "Картриджи"),
        ("электронные системы", "devices", "Электронные системы"),
    )
    for marker, key, label in rules:
        if marker in normalized:
            return key, label
    segments = [segment.strip() for segment in full_path.split("/") if segment.strip()]
    fallback = segments[-2] if len(segments) > 1 else "Прочее"
    return normalize_price_text(fallback), fallback


def to_decimal(value) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        result = Decimal(str(value).replace(",", "."))
    except Exception:
        return None
    return result if result > 0 else None


@dataclass(frozen=True)
class ParsedItem:
    code: str
    name: str
    cash: Decimal
    cashless: Decimal


@dataclass(frozen=True)
class ParsedGroup:
    merge_key: str
    display_name: str
    full_path: str
    category_key: str
    category_name: str
    items: tuple[ParsedItem, ...]


@dataclass(frozen=True)
class ParsedPrice:
    sheet_name: str
    price_date: str | None
    groups: tuple[ParsedGroup, ...]
    action_count: int

    @property
    def item_count(self) -> int:
        return sum(len(group.items) for group in self.groups)


@dataclass(frozen=True)
class GroupSummary:
    callback_id: int
    merge_key: str
    display_name: str
    category_name: str
    search_text: str
    warehouse_counts: dict[str, int]


@dataclass(frozen=True)
class PriceTier:
    cash: Decimal
    cashless: Decimal
    variant_count: int


@dataclass(frozen=True)
class GroupDetails:
    summary: GroupSummary
    tiers: tuple[PriceTier, ...]
    unique_variants: int


@dataclass(frozen=True)
class VariantAvailability:
    name: str
    warehouses: tuple[str, ...]


def parse_price_file(path: Path) -> ParsedPrice:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = None
    buffered = []
    price_date = None
    for number, row in enumerate(rows, 1):
        values = tuple(row)
        buffered.append((number, values))
        first = normalize_price_text(str(values[0] or "")) if values else ""
        if "актуален" in first and len(values) > 1 and values[1]:
            raw_date = values[1]
            price_date = raw_date.isoformat() if hasattr(raw_date, "isoformat") else str(raw_date)
        normalized = [normalize_price_text(str(value or "")) for value in values]
        if (
            any(value == "наименование" for value in normalized)
            and any("50т р нал" in value for value in normalized)
            and any("50т р безнал" in value for value in normalized)
        ):
            header_row = (number, normalized)
            break
        if number >= 30:
            break
    if not header_row:
        workbook.close()
        raise ValueError("Не найдена строка заголовков прайса.")

    _, headers = header_row
    name_idx = headers.index("наименование")
    cash_idx = next(i for i, value in enumerate(headers) if "50т р нал" in value)
    cashless_idx = next(i for i, value in enumerate(headers) if "50т р безнал" in value)
    code_idx = next((i for i, value in enumerate(headers) if value == "код"), 0)

    parsed_groups = []
    current_path = None
    current_items = []
    action_count = 0
    action_groups = {}

    def finish_group() -> None:
        nonlocal current_items
        if not current_path or not current_items:
            current_items = []
            return
        display_name = clean_group_name(current_path)
        category_key, category_name = category_from_path(current_path)
        merge_key = f"{category_key}|{normalize_price_text(display_name)}"
        parsed_groups.append(
            ParsedGroup(
                merge_key, display_name, current_path, category_key, category_name,
                tuple(current_items),
            )
        )
        current_items = []

    for row in rows:
        values = tuple(row)
        first = str(values[0] or "").strip() if values else ""
        name = str(values[name_idx] or "").strip() if len(values) > name_idx else ""
        cash = to_decimal(values[cash_idx] if len(values) > cash_idx else None)
        cashless = to_decimal(values[cashless_idx] if len(values) > cashless_idx else None)
        if first and "/" in first and (not name or cash is None or cashless is None):
            finish_group()
            current_path = first
            continue
        if not name or cash is None or cashless is None or not current_path:
            continue
        if normalize_price_text(name).startswith("акция "):
            action_count += 1
            clean_name = re.sub(r"^\s*АКЦИЯ\s+", "", name, flags=re.IGNORECASE)
            base_name = clean_name.split(" - ", 1)[0].strip() if " - " in clean_name else clean_name
            category_key, category_name = category_from_path(current_path)
            merge_key = f"{category_key}|{normalize_price_text(base_name)}"
            if merge_key not in action_groups:
                action_groups[merge_key] = {
                    "display": base_name,
                    "path": f"{current_path}/{base_name}",
                    "category_key": category_key,
                    "category_name": category_name,
                    "items": [],
                }
            code = str(values[code_idx] or "").strip() if len(values) > code_idx else ""
            action_groups[merge_key]["items"].append(ParsedItem(code, clean_name, cash, cashless))
            continue
        code = str(values[code_idx] or "").strip() if len(values) > code_idx else ""
        current_items.append(ParsedItem(code, name, cash, cashless))
    finish_group()
    workbook.close()
    for merge_key, group in action_groups.items():
        parsed_groups.append(
            ParsedGroup(
                merge_key,
                group["display"],
                group["path"],
                group["category_key"],
                group["category_name"],
                tuple(group["items"]),
            )
        )
    combined = {}
    for group in parsed_groups:
        existing = combined.get(group.merge_key)
        if existing:
            combined[group.merge_key] = ParsedGroup(
                existing.merge_key,
                existing.display_name,
                f"{existing.full_path} {group.full_path}",
                existing.category_key,
                existing.category_name,
                existing.items + group.items,
            )
        else:
            combined[group.merge_key] = group
    parsed_groups = list(combined.values())
    if not parsed_groups or not sum(len(group.items) for group in parsed_groups):
        raise ValueError("В прайсе не найдено товарных групп с корректными ценами.")
    return ParsedPrice(sheet.title, price_date, tuple(parsed_groups), action_count)


class PricesDB:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        return connection

    def _initialize(self) -> None:
        with closing(self._connect()) as connection, connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS price_imports (
                    warehouse TEXT PRIMARY KEY,
                    file_name TEXT NOT NULL,
                    price_date TEXT,
                    item_count INTEGER NOT NULL,
                    group_count INTEGER NOT NULL,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS price_groups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    warehouse TEXT NOT NULL,
                    merge_key TEXT NOT NULL,
                    display_name TEXT NOT NULL,
                    full_path TEXT NOT NULL,
                    category_name TEXT NOT NULL,
                    position INTEGER NOT NULL,
                    UNIQUE(warehouse, merge_key)
                );
                CREATE TABLE IF NOT EXISTS price_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    group_id INTEGER NOT NULL REFERENCES price_groups(id) ON DELETE CASCADE,
                    code TEXT,
                    name TEXT NOT NULL,
                    cash TEXT NOT NULL,
                    cashless TEXT NOT NULL,
                    position INTEGER NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_price_groups_merge_key ON price_groups(merge_key);
                CREATE INDEX IF NOT EXISTS idx_price_items_group_id ON price_items(group_id);
                """
            )

    def replace_warehouse(self, warehouse: str, parsed: ParsedPrice, file_name: str) -> None:
        if warehouse not in WAREHOUSES:
            raise ValueError("Неизвестный склад.")
        with closing(self._connect()) as connection, connection:
            connection.execute("DELETE FROM price_groups WHERE warehouse = ?", (warehouse,))
            for group_position, group in enumerate(parsed.groups):
                cursor = connection.execute(
                    """INSERT INTO price_groups(
                           warehouse, merge_key, display_name, full_path, category_name, position
                       ) VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        warehouse, group.merge_key, group.display_name, group.full_path,
                        group.category_name, group_position,
                    ),
                )
                group_id = cursor.lastrowid
                connection.executemany(
                    """INSERT INTO price_items(group_id, code, name, cash, cashless, position)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    [
                        (group_id, item.code, item.name, str(item.cash), str(item.cashless), position)
                        for position, item in enumerate(group.items)
                    ],
                )
            connection.execute(
                """INSERT INTO price_imports(
                       warehouse, file_name, price_date, item_count, group_count, updated_at
                   ) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                   ON CONFLICT(warehouse) DO UPDATE SET
                       file_name=excluded.file_name, price_date=excluded.price_date,
                       item_count=excluded.item_count, group_count=excluded.group_count,
                       updated_at=CURRENT_TIMESTAMP""",
                (warehouse, file_name, parsed.price_date, parsed.item_count, len(parsed.groups)),
            )

    def import_statuses(self) -> dict[str, sqlite3.Row]:
        with closing(self._connect()) as connection, connection:
            rows = connection.execute("SELECT * FROM price_imports").fetchall()
        return {row["warehouse"]: row for row in rows}

    def group_summaries(self) -> list[GroupSummary]:
        with closing(self._connect()) as connection, connection:
            groups = connection.execute(
                """SELECT g.id, g.merge_key, g.display_name, g.category_name, g.full_path,
                          g.warehouse, COUNT(i.id) AS item_count
                   FROM price_groups g JOIN price_items i ON i.group_id = g.id
                   GROUP BY g.id ORDER BY g.position"""
            ).fetchall()
        merged = {}
        for row in groups:
            key = row["merge_key"]
            if key not in merged:
                merged[key] = {
                    "id": row["id"],
                    "display": row["display_name"], "category": row["category_name"],
                    "search": [], "counts": {},
                }
            merged[key]["search"].append(row["full_path"])
            merged[key]["counts"][row["warehouse"]] = row["item_count"]
        return [
            GroupSummary(value["id"], key, value["display"], value["category"], " ".join(value["search"]), value["counts"])
            for key, value in merged.items()
        ]

    def search_groups(self, query: str, limit: int = 100) -> list[GroupSummary]:
        query_norm = normalize_price_text(query)
        query_tokens = query_norm.split()
        if not query_tokens:
            return []
        category_intent = None
        category_markers = (
            (("жидкост",), "Жидкости"),
            (("ароматизатор", "конструктор"), "Конструкторы и ароматизаторы"),
            (("однораз",), "Одноразовые системы"),
            (("картридж",), "Картриджи"),
            (("электронн",), "Электронные системы"),
        )
        for markers, category in category_markers:
            if any(any(token.startswith(marker) for marker in markers) for token in query_tokens):
                category_intent = category
                break
        ranked = []
        for group in self.group_summaries():
            if category_intent and group.category_name != category_intent:
                continue
            candidate = normalize_price_text(
                f"{group.display_name} {group.category_name} {group.search_text}"
            )
            candidate_tokens = candidate.split()
            scores = []
            for token in query_tokens:
                if token.isdigit():
                    scores.append(1.0 if token in candidate_tokens else 0.0)
                elif token in candidate_tokens:
                    scores.append(1.0)
                elif len(token) >= 5 and token in candidate:
                    scores.append(0.92)
                else:
                    scores.append(max((SequenceMatcher(None, token, word).ratio() for word in candidate_tokens), default=0))
            score = sum(scores) / len(scores)
            token_matches = all(
                value >= (1.0 if token.isdigit() else 0.72 if len(token) <= 4 else 0.58)
                for token, value in zip(query_tokens, scores)
            )
            if token_matches and score >= 0.68:
                exact_bonus = 0.15 if query_norm in normalize_price_text(group.display_name) else 0
                ranked.append((score + exact_bonus, sum(group.warehouse_counts.values()), group))
        ranked.sort(key=lambda value: (value[0], value[1]), reverse=True)
        buckets = defaultdict(list)
        for score, item_count, group in ranked:
            buckets[group.category_name].append((score, item_count, group))
        preferred_categories = (
            "Жидкости",
            "Одноразовые системы",
            "Электронные системы",
            "Картриджи",
            "Конструкторы и ароматизаторы",
        )
        category_order = [category for category in preferred_categories if category in buckets]
        category_order.extend(category for category in buckets if category not in category_order)
        diversified = []
        while any(buckets.values()):
            for category in category_order:
                if buckets[category]:
                    diversified.append(buckets[category].pop(0)[2])
                    if len(diversified) >= limit:
                        return diversified
        return diversified

    def group_details(self, callback_id: int) -> GroupDetails | None:
        summaries = {group.callback_id: group for group in self.group_summaries()}
        summary = summaries.get(callback_id)
        if not summary:
            return None
        merge_key = summary.merge_key
        with closing(self._connect()) as connection, connection:
            rows = connection.execute(
                """SELECT i.code, i.name, i.cash, i.cashless
                   FROM price_items i JOIN price_groups g ON g.id = i.group_id
                   WHERE g.merge_key = ?""",
                (merge_key,),
            ).fetchall()
        tiers = defaultdict(set)
        variants = set()
        for row in rows:
            identity = row["code"] or normalize_price_text(row["name"])
            variants.add(identity)
            tiers[(Decimal(row["cash"]), Decimal(row["cashless"]))].add(identity)
        price_tiers = tuple(
            PriceTier(cash, cashless, len(identities))
            for (cash, cashless), identities in sorted(tiers.items())
        )
        return GroupDetails(summary, price_tiers, len(variants))

    def group_variants(self, callback_id: int) -> tuple[GroupSummary, tuple[VariantAvailability, ...]] | None:
        summaries = {group.callback_id: group for group in self.group_summaries()}
        summary = summaries.get(callback_id)
        if not summary:
            return None
        with closing(self._connect()) as connection, connection:
            rows = connection.execute(
                """SELECT g.warehouse, i.code, i.name
                   FROM price_items i JOIN price_groups g ON g.id = i.group_id
                   WHERE g.merge_key = ? ORDER BY i.position""",
                (summary.merge_key,),
            ).fetchall()
        variants = {}
        for row in rows:
            identity = row["code"] or normalize_price_text(row["name"])
            if identity not in variants:
                variants[identity] = {"name": variant_display_name(row["name"]), "warehouses": set()}
            variants[identity]["warehouses"].add(row["warehouse"])
        result = tuple(
            VariantAvailability(value["name"], tuple(sorted(value["warehouses"])))
            for value in sorted(variants.values(), key=lambda item: normalize_price_text(item["name"]))
        )
        return summary, result

    def selection_availability(
        self, callback_ids: list[int]
    ) -> tuple[list[str], dict[str, set[str]]]:
        summaries = {group.callback_id: group for group in self.group_summaries()}
        merge_keys = list(dict.fromkeys(
            summaries[value].merge_key for value in callback_ids if value in summaries
        ))
        if not merge_keys:
            return [], {}
        placeholders = ",".join("?" for _ in merge_keys)
        with closing(self._connect()) as connection:
            rows = connection.execute(
                f"""SELECT g.warehouse, i.code, i.name
                    FROM price_items i JOIN price_groups g ON g.id = i.group_id
                    WHERE g.merge_key IN ({placeholders})""",
                merge_keys,
            ).fetchall()
        availability: dict[str, set[str]] = defaultdict(set)
        for row in rows:
            identity = row["code"] or normalize_price_text(row["name"])
            availability[identity].add(row["warehouse"])
        return merge_keys, dict(availability)

    def backup_to(self, destination: Path) -> None:
        destination.parent.mkdir(parents=True, exist_ok=True)
        source = self._connect()
        target = sqlite3.connect(destination)
        try:
            source.backup(target)
        finally:
            target.close()
            source.close()


def save_price_source(source: Path, storage_dir: Path, warehouse: str) -> Path | None:
    storage_dir.mkdir(parents=True, exist_ok=True)
    destination = storage_dir / f"{warehouse}.xlsx"
    backup = None
    if destination.exists():
        backup_dir = storage_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup = backup_dir / f"{warehouse}_{timestamp}.xlsx"
        shutil.copy2(destination, backup)
    shutil.copy2(source, destination)
    return backup


def generate_discounted_price(source: Path, destination: Path, percent: int = 10) -> int:
    if percent <= 0 or percent >= 100:
        raise ValueError("Скидка должна быть от 1 до 99 процентов.")
    workbook = load_workbook(source, data_only=False)
    sheet = workbook.active
    header_row = None
    headers = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=30), 1):
        normalized = [normalize_price_text(str(cell.value or "")) for cell in row]
        if (
            any(value == "наименование" for value in normalized)
            and any("50т р нал" in value for value in normalized)
            and any("50т р безнал" in value for value in normalized)
        ):
            header_row, headers = row_number, normalized
            break
    if not header_row or headers is None:
        workbook.close()
        raise ValueError("Не найдена строка заголовков прайса.")
    name_idx = headers.index("наименование") + 1
    cash_idx = next(i for i, value in enumerate(headers, 1) if "50т р нал" in value)
    cashless_idx = next(i for i, value in enumerate(headers, 1) if "50т р безнал" in value)
    sheet.cell(header_row, cash_idx).value = f"от 50т.р. нал — скидка {percent}%"
    sheet.cell(header_row, cashless_idx).value = f"от 50т.р. безнал — скидка {percent}%"
    multiplier = Decimal(100 - percent) / Decimal(100)
    changed = 0
    for row_number in range(header_row + 1, sheet.max_row + 1):
        if not sheet.cell(row_number, name_idx).value:
            continue
        row_changed = False
        for column in (cash_idx, cashless_idx):
            cell = sheet.cell(row_number, column)
            original = to_decimal(cell.value)
            if original is None:
                continue
            discounted_price = (original * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            cell.value = float(discounted_price)
            cell.number_format = "0.00"
            row_changed = True
        if row_changed:
            changed += 1
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()
    return changed


def generate_selected_price(
    source: Path,
    destination: Path,
    merge_keys: list[str],
    availability: dict[str, set[str]],
    discount: int = 0,
) -> int:
    """Filter a native warehouse price while preserving its original workbook design."""
    if discount not in (0, 10):
        raise ValueError("Поддерживаются базовые цены или скидка 10%.")
    if not merge_keys:
        raise ValueError("В подборке нет доступных товарных групп.")
    selected_keys = set(merge_keys)
    workbook = load_workbook(source, data_only=False)
    sheet = workbook.active
    header_row = None
    headers = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=30), 1):
        normalized = [normalize_price_text(str(cell.value or "")) for cell in row]
        if (
            any(value == "наименование" for value in normalized)
            and any("50т р нал" in value for value in normalized)
            and any("50т р безнал" in value for value in normalized)
        ):
            header_row, headers = row_number, normalized
            break
    if not header_row or headers is None:
        workbook.close()
        raise ValueError("Не найдена строка заголовков прайса.")
    name_col = headers.index("наименование") + 1
    cash_col = next(i for i, value in enumerate(headers, 1) if "50т р нал" in value)
    cashless_col = next(i for i, value in enumerate(headers, 1) if "50т р безнал" in value)

    keep_rows = [True] * (sheet.max_row + 1)
    group_rows: list[int] = []
    selected_group_rows: set[int] = set()
    selected_item_rows: set[int] = set()
    selected_identities: dict[int, str] = {}
    current_path = None
    current_group_row = None
    first_group_row = None
    last_item_row = None
    for row_number in range(header_row + 1, sheet.max_row + 1):
        first = str(sheet.cell(row_number, 1).value or "").strip()
        name = str(sheet.cell(row_number, name_col).value or "").strip()
        cash = to_decimal(sheet.cell(row_number, cash_col).value)
        cashless = to_decimal(sheet.cell(row_number, cashless_col).value)
        if first and "/" in first and (not name or cash is None or cashless is None):
            current_path = first
            current_group_row = row_number
            group_rows.append(row_number)
            first_group_row = first_group_row or row_number
            continue
        if not current_path or not name or cash is None or cashless is None:
            continue
        last_item_row = row_number
        category_key, _ = category_from_path(current_path)
        if normalize_price_text(name).startswith("акция "):
            clean_name = re.sub(r"^\s*АКЦИЯ\s+", "", name, flags=re.IGNORECASE)
            base_name = clean_name.split(" - ", 1)[0].strip() if " - " in clean_name else clean_name
            row_key = f"{category_key}|{normalize_price_text(base_name)}"
        else:
            row_key = f"{category_key}|{normalize_price_text(clean_group_name(current_path))}"
        if row_key in selected_keys:
            selected_item_rows.add(row_number)
            code = str(sheet.cell(row_number, 1).value or "").strip()
            comparable_name = re.sub(r"^\s*АКЦИЯ\s+", "", name, flags=re.IGNORECASE)
            selected_identities[row_number] = code or normalize_price_text(comparable_name)
            if current_group_row is not None:
                selected_group_rows.add(current_group_row)

    if not selected_item_rows or first_group_row is None or last_item_row is None:
        workbook.close()
        raise ValueError("Выбранные товары отсутствуют в исходном прайсе.")
    for row_number in range(first_group_row, last_item_row + 1):
        keep_rows[row_number] = row_number in selected_item_rows or row_number in selected_group_rows

    original_max_row = sheet.max_row
    formulas = {
        (cell.row, cell.column): cell.value
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=") and keep_rows[cell.row]
    }
    original_dimensions = {index: copy(dimension) for index, dimension in sheet.row_dimensions.items()}
    merged_ranges = [copy(cell_range) for cell_range in sheet.merged_cells.ranges]
    anchored_objects = [*getattr(sheet, "_images", []), *getattr(sheet, "_charts", [])]
    for cell_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(cell_range))

    row_map = {}
    new_row = 0
    for old_row in range(1, original_max_row + 1):
        if keep_rows[old_row]:
            new_row += 1
            row_map[old_row] = new_row
    remove_ranges = []
    range_start = None
    for row_number in range(1, original_max_row + 2):
        remove = row_number <= original_max_row and not keep_rows[row_number]
        if remove and range_start is None:
            range_start = row_number
        elif not remove and range_start is not None:
            remove_ranges.append((range_start, row_number - range_start))
            range_start = None
    for start, amount in reversed(remove_ranges):
        sheet.delete_rows(start, amount)

    removed_anchors = []
    for anchored in anchored_objects:
        anchor = getattr(anchored, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        original_row = marker.row + 1
        if original_row in row_map:
            marker.row = row_map[original_row] - 1
        else:
            removed_anchors.append(anchored)
            continue
        end_marker = getattr(anchor, "to", None)
        if end_marker is not None:
            original_end_row = end_marker.row + 1
            if original_end_row in row_map:
                end_marker.row = row_map[original_end_row] - 1
    sheet._images = [value for value in getattr(sheet, "_images", []) if value not in removed_anchors]
    sheet._charts = [value for value in getattr(sheet, "_charts", []) if value not in removed_anchors]

    sheet.row_dimensions.clear()
    for old_row, dimension in original_dimensions.items():
        if old_row not in row_map:
            continue
        mapped = copy(dimension)
        mapped.index = row_map[old_row]
        sheet.row_dimensions[row_map[old_row]] = mapped
    for cell_range in merged_ranges:
        mapped_rows = [row_map.get(row) for row in range(cell_range.min_row, cell_range.max_row + 1)]
        if not mapped_rows or any(row is None for row in mapped_rows):
            continue
        if mapped_rows != list(range(mapped_rows[0], mapped_rows[-1] + 1)):
            continue
        sheet.merge_cells(
            start_row=mapped_rows[0], start_column=cell_range.min_col,
            end_row=mapped_rows[-1], end_column=cell_range.max_col,
        )
    for (old_row, column), formula in formulas.items():
        mapped_row = row_map[old_row]
        old_coordinate = f"{get_column_letter(column)}{old_row}"
        new_coordinate = sheet.cell(mapped_row, column).coordinate
        try:
            translated = Translator(formula, origin=old_coordinate).translate_formula(new_coordinate)
        except Exception:
            translated = formula
        sheet.cell(mapped_row, column).value = translated

    availability_columns = {"center": 9, "west": 10, "ural": 11}
    for warehouse, column in availability_columns.items():
        header = sheet.cell(header_row, column)
        header.value = WAREHOUSES[warehouse]
        header._style = copy(sheet.cell(header_row, cash_col)._style)
        header.alignment = copy(sheet.cell(header_row, cash_col).alignment)
        column_letter = get_column_letter(column)
        current_width = sheet.column_dimensions[column_letter].width or 0
        sheet.column_dimensions[column_letter].width = max(current_width, 20)
    for old_row, identity in selected_identities.items():
        mapped_row = row_map[old_row]
        warehouses = availability.get(identity, set())
        for warehouse, column in availability_columns.items():
            cell = sheet.cell(mapped_row, column)
            cell._style = copy(sheet.cell(mapped_row, cash_col)._style)
            cell.alignment = copy(sheet.cell(mapped_row, cash_col).alignment)
            cell.value = "✅" if warehouse in warehouses else "❌"
    sheet.freeze_panes = "A9"

    if discount:
        sheet.cell(header_row, cash_col).value = f"от 50т.р. нал — скидка {discount}%"
        sheet.cell(header_row, cashless_col).value = f"от 50т.р. безнал — скидка {discount}%"
        multiplier = Decimal(100 - discount) / Decimal(100)
        for old_row in selected_item_rows:
            mapped_row = row_map[old_row]
            for column in (cash_col, cashless_col):
                cell = sheet.cell(mapped_row, column)
                original = to_decimal(cell.value)
                if original is not None:
                    cell.value = float((original * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    cell.number_format = "0.00"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()
    restore_native_drawings(source, destination)
    return len(selected_item_rows)


def restore_native_drawings(source: Path, destination: Path) -> None:
    """Restore drawing parts unsupported by openpyxl (notably absolute anchors)."""
    with ZipFile(source) as source_zip, ZipFile(destination) as output_zip:
        source_names = set(source_zip.namelist())
        source_sheet = source_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")
        output_sheet = output_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")
        drawing_tags = re.findall(
            r"<(?:drawing|legacyDrawing|legacyDrawingHF|picture)\b[^>]*/>", source_sheet
        )
        if drawing_tags and not any(tag.split()[0] in output_sheet for tag in drawing_tags):
            if "xmlns:r=" not in output_sheet:
                output_sheet = output_sheet.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1,
                )
            output_sheet = output_sheet.replace("</worksheet>", "".join(drawing_tags) + "</worksheet>")

        content_types = ElementTree.fromstring(output_zip.read("[Content_Types].xml"))
        source_types = ElementTree.fromstring(source_zip.read("[Content_Types].xml"))
        existing_types = {(child.tag, tuple(sorted(child.attrib.items()))) for child in content_types}
        for child in source_types:
            signature = (child.tag, tuple(sorted(child.attrib.items())))
            if signature not in existing_types and (
                "drawing" in child.attrib.get("PartName", "")
                or child.attrib.get("Extension", "").casefold() in {"jpeg", "jpg", "png", "gif", "svg"}
            ):
                content_types.append(copy(child))

        with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=destination.parent, delete=False) as temp_file:
            replacement = Path(temp_file.name)
        try:
            with ZipFile(replacement, "w", ZIP_DEFLATED) as rebuilt:
                for item in output_zip.infolist():
                    name = item.filename
                    if name in {"xl/worksheets/sheet1.xml", "[Content_Types].xml"}:
                        continue
                    if name == "xl/worksheets/_rels/sheet1.xml.rels" and name in source_names:
                        continue
                    if name.startswith(("xl/drawings/", "xl/media/")) and name in source_names:
                        continue
                    rebuilt.writestr(item, output_zip.read(name))
                rebuilt.writestr("xl/worksheets/sheet1.xml", output_sheet.encode("utf-8"))
                rebuilt.writestr(
                    "[Content_Types].xml",
                    ElementTree.tostring(content_types, encoding="utf-8", xml_declaration=True),
                )
                sheet_relationships = "xl/worksheets/_rels/sheet1.xml.rels"
                if sheet_relationships in source_names:
                    rebuilt.writestr(sheet_relationships, source_zip.read(sheet_relationships))
                for name in source_names:
                    if name.startswith(("xl/drawings/", "xl/media/")):
                        rebuilt.writestr(name, source_zip.read(name))
            shutil.move(replacement, destination)
        finally:
            replacement.unlink(missing_ok=True)
