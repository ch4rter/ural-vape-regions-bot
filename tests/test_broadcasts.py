from openpyxl import Workbook, load_workbook

from bot import build_broadcast_report, build_chats_excel, parse_audience_excel
import bot
from materials_db import MaterialsDB


def test_audience_excel_parsing(tmp_path):
    path = tmp_path / "audience.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Клиент", "Chat ID"])
    sheet.append(["Первый", -100123])
    sheet.append(["Дубль", -100123])
    sheet.append(["Второй", "-100456"])
    sheet.append(["Ошибка", "abc"])
    workbook.save(path)

    chat_ids, duplicates, invalid = parse_audience_excel(path)
    assert chat_ids == [-100123, -100456]
    assert duplicates == 1
    assert invalid == 1


def test_chat_export_and_broadcast_report(tmp_path):
    previous = getattr(bot, "materials_db", None)
    bot.materials_db = MaterialsDB(tmp_path / "materials.sqlite3")
    bot.materials_db.upsert_client_chat(-100123, "Клиент", "supergroup", True)
    try:
        chats_path = tmp_path / "chats.xlsx"
        build_chats_excel(chats_path)
        workbook = load_workbook(chats_path, data_only=True)
        assert workbook.active["A2"].value == "Клиент"
        assert workbook.active["B2"].value == -100123
        workbook.close()

        report_path = tmp_path / "report.xlsx"
        build_broadcast_report(report_path, [{
            "title": "Клиент", "chat_id": -100123,
            "status": "Отправлено", "error": "",
        }])
        report = load_workbook(report_path, data_only=True)
        assert report.active["C2"].value == "Отправлено"
        report.close()
    finally:
        if previous is not None:
            bot.materials_db = previous
