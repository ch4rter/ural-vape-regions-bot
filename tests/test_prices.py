from decimal import Decimal

from openpyxl import Workbook, load_workbook

from bot import (
    discounted,
    format_combined_report_notification,
    format_price_item,
    format_price_report,
    manager_html,
    money,
    variant_word,
)
from prices_db import (
    ParsedGroup,
    ParsedItem,
    ParsedPrice,
    PricesDB,
    clean_group_name,
    generate_discounted_price,
    generate_change_report_excel,
    generate_selected_price,
    parse_price_file,
)


def make_price(path, suffix="", include_action=True):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ПРАЙС-ЛИСТ"])
    for _ in range(6):
        sheet.append([])
    sheet.append(["Код", "Наименование", "Ед.изм.", "от 50т.р. нал", "от 50т.р. безнал"])
    sheet.append(["ЭС/3.Жидкости/Производитель OGGO/VLIQ OGGO BALANCE 20"])
    sheet.append(["001", f"VLIQ OGGO BALANCE Манго{suffix}", "шт", 235, 259])
    if include_action:
        sheet.append(["002", "АКЦИЯ VLIQ OGGO BALANCE Арбуз", "шт", 100, 110])
    sheet.append(["ЭС/3.Жидкости/Производитель OGGO/VLIQ OGGO BALANCE 20"])
    sheet.append(["003", f"VLIQ OGGO BALANCE Вишня{suffix}", "шт", 247, 270])
    sheet.merge_cells("A2:E2")
    sheet["A2"] = "Минимальный заказ — 50 000 ₽"
    sheet["F10"] = "=D10*G10"
    sheet["F13"] = "=D13*G13"
    workbook.save(path)


def test_price_parser_uses_group_rows_and_ignores_actions(tmp_path):
    path = tmp_path / "price.xlsx"
    make_price(path)
    parsed = parse_price_file(path)

    assert len(parsed.groups) == 2
    regular = next(group for group in parsed.groups if group.display_name == "VLIQ OGGO BALANCE 20")
    action = next(group for group in parsed.groups if group.display_name == "VLIQ OGGO BALANCE Арбуз")
    assert regular.category_name == "Жидкости"
    assert len(regular.items) == 2
    assert action.items[0].name == "VLIQ OGGO BALANCE Арбуз"
    assert parsed.action_count == 1


def test_warehouse_replacement_search_and_aggregation(tmp_path):
    center = tmp_path / "center.xlsx"
    west = tmp_path / "west.xlsx"
    make_price(center, include_action=False)
    make_price(west, suffix=" West", include_action=False)
    db = PricesDB(tmp_path / "prices.sqlite3")
    db.replace_warehouse("center", parse_price_file(center), center.name)
    db.replace_warehouse("west", parse_price_file(west), west.name)

    results = db.search_groups("ogo vlq balance")
    assert len(results) == 1
    assert results[0].warehouse_counts == {"center": 2, "west": 2}
    details = db.group_details(results[0].callback_id)
    assert len(details.tiers) == 2
    assert details.unique_variants == 2
    selected_keys, availability = db.selection_availability([results[0].callback_id])
    assert selected_keys == [results[0].merge_key]
    assert availability["001"] == {"center", "west"}

    replacement = tmp_path / "replacement.xlsx"
    make_price(replacement, suffix=" New", include_action=False)
    db.replace_warehouse("center", parse_price_file(replacement), replacement.name)
    result = db.search_groups("VLIQ BALANCE")[0]
    assert result.warehouse_counts == {"west": 2, "center": 2}

    liquid_results = db.search_groups("жидкость OGGO")
    assert liquid_results
    assert all(group.category_name == "Жидкости" for group in liquid_results)


def test_prices_and_group_name_formatting():
    assert clean_group_name("ЭС/Одноразовые/1. Д Vaporesso Dojo 12000") == "Vaporesso Dojo 12000"
    assert clean_group_name("ЭС/Расходники/Vaporesso/Расходники") == "Vaporesso Расходники"
    assert clean_group_name("ЭС/Расходники/GeekVape/Расходники") == "GeekVape Расходники"
    assert clean_group_name("ЭС/Жидкости/Производитель OGGO X ELFLIQ/Ice 20mg") == "OGGO X ELFLIQ Ice 20mg"
    assert clean_group_name("ЭС/Жидкости/Производитель OGGO/Oggo Acid") == "Oggo Acid"
    assert money(discounted(Decimal("235"), 5)) == "223,25"
    assert variant_word(1) == "вариант"
    assert variant_word(2) == "варианта"
    assert variant_word(15) == "вариантов"
    assert manager_html("Андрей") == '<b><a href="https://t.me/shmidtuv">Андрей</a></b>'
    assert manager_html("Другой") == "<b>Другой</b>"


def test_discounted_excel_changes_copy_but_not_original(tmp_path):
    source = tmp_path / "base.xlsx"
    destination = tmp_path / "discount.xlsx"
    make_price(source)

    changed = generate_discounted_price(source, destination, 10)
    assert changed == 3

    base = load_workbook(source, data_only=True)
    discounted_book = load_workbook(destination, data_only=True)
    assert base.active["D10"].value == 235
    assert discounted_book.active["D8"].value == "от 50т.р. нал — скидка 10%"
    assert discounted_book.active["E8"].value == "от 50т.р. безнал — скидка 10%"
    assert discounted_book.active["D10"].value == 211.5
    assert discounted_book.active["E10"].value == 233.1
    assert discounted_book.active["D11"].value == 90
    base.close()
    discounted_book.close()


def test_parser_keeps_brand_context_for_generic_and_short_group_names(tmp_path):
    path = tmp_path / "brands.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for _ in range(7):
        sheet.append([])
    sheet.append(["Код", "Наименование", "Ед.изм.", "от 50т.р. нал", "от 50т.р. безнал"])
    sheet.append(["ЭС/5.Расходники/Vaporesso/Расходники"])
    sheet.append(["V1", "Испаритель Vaporesso", "шт", 100, 110])
    sheet.append(["ЭС/5.Расходники/GeekVape/Расходники"])
    sheet.append(["G1", "Испаритель GeekVape", "шт", 120, 130])
    sheet.append(["ЭС/3.Жидкости/Производитель OGGO X ELFLIQ/Ice 20mg"])
    sheet.append(["O1", "OGGO X ELFLIQ Ice Манго", "шт", 230, 250])
    workbook.save(path)

    parsed = parse_price_file(path)
    names = {group.display_name for group in parsed.groups}
    keys = {group.merge_key for group in parsed.groups}
    assert names == {"Vaporesso Расходники", "GeekVape Расходники", "OGGO X ELFLIQ Ice 20mg"}
    assert "consumables|vaporesso расходники" in keys
    assert "consumables|geekvape расходники" in keys
    assert "liquids|oggo x elfliq ice 20mg" in keys


def test_specific_position_search_uses_name_not_product_code(tmp_path):
    db = PricesDB(tmp_path / "prices.sqlite3")
    group = ParsedGroup(
        "consumables|vaporesso расходники", "Vaporesso Расходники",
        "ЭС/5.Расходники/Vaporesso/Расходники", "consumables", "Расходники",
        (
            ParsedItem("027881", "Картридж Vaporesso XROS (2мл) - 0.6 ohm COREX 3.0", Decimal("532"), Decimal("549")),
            ParsedItem("027180", "Картридж Vaporesso XROS (3мл) - 0.6 ohm", Decimal("556"), Decimal("570")),
        ),
    )
    db.replace_warehouse("center", ParsedPrice("Прайс", None, (group,), 0), "center.xlsx")
    west_group = ParsedGroup(
        group.merge_key, group.display_name, group.full_path, group.category_key, group.category_name,
        (ParsedItem("027881", group.items[0].name, Decimal("532"), Decimal("549")),),
    )
    db.replace_warehouse("west", ParsedPrice("Прайс", None, (west_group,), 0), "west.xlsx")

    results = db.search_items("vaporesso xros 0.6 2мл")
    assert len(results) == 1
    assert "(2мл)" in results[0].name
    assert results[0].warehouse_prices == {
        "center": (Decimal("532"), Decimal("549")),
        "west": (Decimal("532"), Decimal("549")),
    }
    assert db.search_items("027881") == []

    card = format_price_item(results[0])
    assert "Москва · Санкт-Петербург" in card
    assert "−15%" in card
    assert "027881" not in card

    _, variants = db.group_variants(db.search_groups("vaporesso")[0].callback_id)
    two_ml = next(item for item in variants if "2мл" in item.name)
    assert two_ml.warehouse_prices["center"] == (Decimal("532"), Decimal("549"))


def test_selected_price_contains_only_chosen_groups_and_discount(tmp_path):
    source = tmp_path / "center.xlsx"
    make_price(source)
    db = PricesDB(tmp_path / "prices.sqlite3")
    db.replace_warehouse("center", parse_price_file(source), source.name)
    selected = db.search_groups("VLIQ OGGO BALANCE 20")[0]
    destination = tmp_path / "selection.xlsx"

    merge_keys, availability = db.selection_availability([selected.callback_id])
    changed = generate_selected_price(source, destination, merge_keys, availability, 10)

    assert changed == 2
    workbook = load_workbook(destination, data_only=True)
    sheet = workbook.active
    values = [cell.value for cell in sheet["B"]]
    assert "VLIQ OGGO BALANCE Манго" in values
    assert "VLIQ OGGO BALANCE Вишня" in values
    assert "VLIQ OGGO BALANCE Арбуз" not in values
    mango_row = values.index("VLIQ OGGO BALANCE Манго") + 1
    assert sheet.cell(mango_row, 4).value == 211.5
    assert sheet.cell(mango_row, 5).value == 233.1
    assert [sheet.cell(8, column).value for column in (9, 10, 11)] == [
        "Москва", "Санкт-Петербург", "Челябинск"
    ]
    assert sheet.cell(mango_row, 9).value == "✅"
    assert sheet.cell(mango_row, 10).value == "❌"
    assert sheet.freeze_panes == "A9"
    assert "A2:E2" in {str(value) for value in sheet.merged_cells.ranges}
    cherry_row = values.index("VLIQ OGGO BALANCE Вишня") + 1
    workbook.close()
    formulas = load_workbook(destination, data_only=False)
    assert formulas.active.cell(cherry_row, 6).value == f"=D{cherry_row}*G{cherry_row}"
    formulas.close()


def test_latest_price_change_report_replaces_previous_report(tmp_path):
    source = tmp_path / "old.xlsx"
    make_price(source, include_action=False)
    db = PricesDB(tmp_path / "prices.sqlite3")
    old = parse_price_file(source)
    db.replace_warehouse("center", old, source.name)
    original_group = old.groups[0]
    new = ParsedPrice(
        "Sheet", "2026-07-22", (
            ParsedGroup(
                original_group.merge_key, original_group.display_name, original_group.full_path,
                original_group.category_key, original_group.category_name,
                (
                    ParsedItem("001", "VLIQ OGGO BALANCE Манго", Decimal("250"), Decimal("275")),
                    ParsedItem("004", "VLIQ OGGO BALANCE Киви", Decimal("240"), Decimal("265")),
                ),
            ),
        ), 0,
    )

    report = db.build_change_report("center", new, "new.xlsx")
    assert report is not None
    formatted = format_price_report(report)
    assert isinstance(formatted, str)
    assert "Изменения прайса" in formatted
    assert "Появилось: <b>1</b>" in formatted
    combined = {
        warehouse: {**report, "warehouse": warehouse}
        for warehouse in ("center", "west", "ural")
    }
    notification = format_combined_report_notification(combined)
    assert "Прайсы обновлены" in notification
    assert "Итого по трём складам" in notification
    assert "Появилось: <b>3</b>" in notification
    assert [item["code"] for item in report["added"]] == ["004"]
    assert [item["code"] for item in report["removed"]] == ["003"]
    assert [item["code"] for item in report["price_changes"]] == ["001"]
    db.save_latest_report(report)
    assert db.latest_report("center")["current_file"] == "new.xlsx"
    assert db.report_broadcast_sent("daily-signature") is False
    db.mark_report_broadcast_sent("daily-signature")
    assert db.report_broadcast_sent("daily-signature") is True

    destination = tmp_path / "changes.xlsx"
    generate_change_report_excel(report, destination)
    workbook = load_workbook(destination, data_only=True)
    assert workbook["Появилось"]["A2"].value == "004"
    assert workbook["Закончилось"]["A2"].value == "003"
    assert workbook["Изменились цены"]["A2"].value == "001"
    workbook.close()
