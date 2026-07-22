import asyncio
import html
import logging
import os
import re
import shutil
import sqlite3
import tempfile
import zipfile
from typing import Any, Awaitable, Callable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from itertools import chain
from pathlib import Path

from aiogram import BaseMiddleware, Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramRetryAfter
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, ChatMemberUpdated, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, Message
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from materials_db import Material, MaterialsDB
from prices_db import (
    WAREHOUSES,
    GroupDetails,
    PricesDB,
    generate_change_report_excel,
    generate_discounted_price,
    generate_selected_price,
    normalize_price_text,
    parse_price_file,
    save_price_source,
)


BASE_DIR = Path(__file__).resolve().parent
RESULTS_PER_PAGE = 10
PRICE_VARIANTS_PER_PAGE = 6
PRICE_ITEMS_PER_PAGE = 6
PRICE_GROUPS_PER_PAGE = 8
MAX_SELECTED_PRICE_GROUPS = 50
SERVICE_CHAT_ID = -5565597780
router = Router()
catalog: "Catalog"
materials_db: MaterialsDB
prices_db: PricesDB
admin_ids: set[int] = set()
active_excel_path: Path
managed_excel_path: Path
price_storage_path: Path
price_updates_in_progress: set[str] = set()
broadcasts_in_progress: set[int] = set()
PRICE_MESSAGE_SETTING = "price_command_message"
DEFAULT_PRICE_COMMAND_MESSAGE = (
    "📄 <b>Актуальные прайсы</b>\n\n"
    "<b>Центр</b> — Москва\n"
    "<b>Урал</b> — Челябинск\n"
    "<b>Запад</b> — Санкт-Петербург\n\n"
    "Ниже представлены последние загруженные версии с базовыми ценами."
)

MANAGER_LINKS = {
    "валера": "uvvalera",
    "андрей": "shmidtuv",
    "матвей": "ural_vape",
    "евгений": "evgenuralv",
}


class AppState(StatesGroup):
    region_search = State()
    price_search = State()


class AdminState(StatesGroup):
    product_name = State()
    product_rename = State()
    section_name = State()
    section_rename = State()
    material_upload = State()
    excel_upload = State()
    excel_confirmation = State()
    price_upload = State()
    price_confirmation = State()
    access_user = State()
    price_message = State()


class BroadcastState(StatesGroup):
    audience_upload = State()
    content_upload = State()
    ready = State()


class WaitState(StatesGroup):
    new_client = State()
    new_query = State()
    new_comment = State()
    edit_query = State()
    edit_comment = State()


class AccessMiddleware(BaseMiddleware):
    async def __call__(
        self,
        handler: Callable[[Any, dict[str, Any]], Awaitable[Any]],
        event: Message | CallbackQuery,
        data: dict[str, Any],
    ) -> Any:
        user = event.from_user
        if getattr(event, "chat", None) and event.chat.type != "private":
            text = event.text or "" if isinstance(event, Message) else ""
            is_price_command = bool(re.match(r"^/прайс(?:@\w+)?\s*$", text, re.IGNORECASE))
            if is_price_command:
                return await handler(event, data)
            is_wait_command = bool(re.match(r"^/wait(?:@\w+)?(?:\s|$)", text, re.IGNORECASE))
            if is_wait_command and user and (
                is_admin(user.id) or materials_db.authorize_user(user.id, user.username)
            ):
                return await handler(event, data)
            return None
        if isinstance(event, CallbackQuery) and event.message and event.message.chat.type != "private":
            return None
        if not user or is_admin(user.id) or materials_db.authorize_user(user.id, user.username):
            return await handler(event, data)
        if isinstance(event, CallbackQuery):
            await event.answer("Доступ к боту не предоставлен.", show_alert=True)
        else:
            username = f"@{user.username}" if user.username else "не указан"
            await event.answer(
                "🔒 <b>Доступ ограничен</b>\n\n"
                "Бот предназначен для сотрудников компании. Передайте администратору один из идентификаторов:\n\n"
                f"• Telegram ID: <code>{user.id}</code>\n"
                f"• Username: <code>{html.escape(username)}</code>"
            )
        return None


def normalize(value: str) -> str:
    value = value.casefold().replace("ё", "е")
    value = re.sub(r"[\s\-–—_,.;:()]+", " ", value)
    return value.strip()


def clean_client_title(value: str) -> str:
    value = re.sub(r"\bURAL\s*VAPE\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s*[|/\\·•—–-]+\s*", " · ", value)
    value = re.sub(r"(?:\s*·\s*){2,}", " · ", value).strip(" ·|/\\—–-")
    return re.sub(r"\s+", " ", value).strip() or "Клиентский чат"


@dataclass(frozen=True)
class Entry:
    name: str
    manager: str
    location: str = ""


class Catalog:
    def __init__(self, path: Path):
        self.path = path
        self.entries: list[Entry] = []
        self.by_name: dict[str, list[int]] = {}
        self.reload()

    def reload(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"Не найден файл {self.path.name}.")
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            workbook.close()
            raise ValueError("Excel-файл пуст.")

        headers = [normalize(str(cell or "")) for cell in first_row]
        manager_idx = next((i for i, value in enumerate(headers) if "менеджер" in value), None)
        territory_idx = next((i for i, value in enumerate(headers) if "территор" in value), None)
        name_idx = next(
            (i for i, value in enumerate(headers) if "назван" in value or "регион" in value), None
        )
        location_idx = next((i for i, value in enumerate(headers) if "местополож" in value), None)
        has_header = manager_idx is not None and (territory_idx is not None or name_idx is not None)
        if has_header:
            name_idx = territory_idx if territory_idx is not None else name_idx
            data_rows = rows
            start_row = 2
        else:
            name_idx, manager_idx, location_idx = 0, 1, None
            data_rows = chain([first_row], rows)
            start_row = 1

        entries: list[Entry] = []
        by_name: dict[str, list[int]] = {}
        for row_number, row in enumerate(data_rows, start=start_row):
            name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] is not None else ""
            manager = (
                str(row[manager_idx]).strip()
                if len(row) > manager_idx and row[manager_idx] is not None else ""
            )
            location = (
                str(row[location_idx]).strip()
                if location_idx is not None and len(row) > location_idx and row[location_idx] is not None
                else ""
            )
            if not name and not manager:
                continue
            if not name or not manager:
                logging.warning("Пропущена неполная строка Excel: %s", row_number)
                continue
            index = len(entries)
            entries.append(Entry(name, manager, location))
            by_name.setdefault(normalize(name), []).append(index)
        workbook.close()
        if not entries:
            raise ValueError("В Excel нет заполненных записей.")
        self.entries, self.by_name = entries, by_name
        logging.info("Загружено строк из Excel: %s", len(entries))

    def exact(self, query: str) -> list[Entry]:
        return [self.entries[i] for i in self.by_name.get(normalize(query), [])]

    def suggestions(self, query: str, limit: int = 5) -> list[int]:
        query_norm = normalize(query)
        if not query_norm:
            return []
        ranked = []
        for name_norm, indexes in self.by_name.items():
            ratio = SequenceMatcher(None, query_norm, name_norm).ratio()
            contains = query_norm in name_norm or name_norm in query_norm
            if ratio >= 0.58 or contains:
                ranked.append((ratio + (0.15 if contains else 0), indexes[0]))
        ranked.sort(reverse=True)
        return [index for _, index in ranked[:limit]]


def validate_excel(path: Path) -> tuple[Catalog, tuple[str, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    sheet_name = sheet.title
    workbook.close()
    if not first_row:
        raise ValueError("Excel-файл пуст.")
    headers = tuple(str(value or "").strip() for value in first_row)
    normalized = [normalize(value) for value in headers]
    required = {
        "местоположение": any("местополож" in value for value in normalized),
        "территория": any("территор" in value for value in normalized),
        "менеджер": any("менеджер" in value for value in normalized),
    }
    missing = [name for name, present in required.items() if not present]
    if missing:
        raise ValueError(f"Не найдены обязательные колонки: {', '.join(missing)}.")
    return Catalog(path), (sheet_name, *headers)


def is_admin(user_id: int | None) -> bool:
    return user_id is not None and user_id in admin_ids


def is_junior_admin(user_id: int | None, username: str | None = None) -> bool:
    return bool(user_id and materials_db.user_role(user_id, username) == "junior_admin")


def can_broadcast(user_id: int | None, username: str | None = None) -> bool:
    return is_admin(user_id) or is_junior_admin(user_id, username)


def can_manage_price_message(user_id: int | None, username: str | None = None) -> bool:
    return is_admin(user_id) or is_junior_admin(user_id, username)


def button_grid(buttons: list[InlineKeyboardButton], columns: int = 2) -> list[list[InlineKeyboardButton]]:
    return [buttons[index : index + columns] for index in range(0, len(buttons), columns)]


def compact_nav(
    back_data: str | None = None,
    *,
    forward_data: str | None = None,
    search_data: str | None = None,
    home: bool = True,
) -> list[InlineKeyboardButton]:
    buttons = []
    if back_data:
        buttons.append(InlineKeyboardButton(text="⬅️", callback_data=back_data))
    if forward_data:
        buttons.append(InlineKeyboardButton(text="➡️", callback_data=forward_data))
    if search_data:
        buttons.append(InlineKeyboardButton(text="🔎", callback_data=search_data))
    if home:
        buttons.append(InlineKeyboardButton(text="🏠", callback_data="main:menu"))
    return buttons


def main_menu(user_id: int | None) -> InlineKeyboardMarkup:
    buttons = [
        InlineKeyboardButton(text="🔎 Менеджеры", callback_data="main:region"),
        InlineKeyboardButton(text="💰 Цены", callback_data="main:prices"),
        InlineKeyboardButton(text="📄 Прайсы", callback_data="main:price_files"),
        InlineKeyboardButton(text="📊 Изменения", callback_data="main:price_reports"),
        InlineKeyboardButton(text="🗃 База данных", callback_data="main:database"),
        InlineKeyboardButton(text="🔔 Ожидания", callback_data="main:waitlist"),
    ]
    rows = button_grid(buttons)
    extra = []
    if can_broadcast(user_id):
        extra.append(InlineKeyboardButton(text="📣 Рассылки", callback_data="main:broadcasts"))
    if can_manage_price_message(user_id):
        extra.append(InlineKeyboardButton(text="⚙️ Управление", callback_data="main:admin"))
    rows.extend(button_grid(extra))
    return InlineKeyboardMarkup(inline_keyboard=rows)


def back_main() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🏠", callback_data="main:menu")]])


async def edit_or_answer(
    message: Message, text: str, reply_markup: InlineKeyboardMarkup | None = None
) -> None:
    try:
        await message.edit_text(text, reply_markup=reply_markup)
    except Exception:
        logging.exception("Не удалось изменить служебное сообщение, отправляю новое")
        await message.answer(text, reply_markup=reply_markup)


async def show_main(target: Message, user_id: int | None, *, edit: bool = False) -> None:
    text = (
        "👋 <b>Добро пожаловать!</b>\n\n"
        "Здесь можно найти ответственного менеджера, проверить цены и наличие, "
        "получить рабочие материалы или открыть свой лист ожидания.\n\n"
        "Выберите нужный раздел:"
    )
    if edit:
        await target.edit_text(text, reply_markup=main_menu(user_id))
    else:
        await target.answer(text, reply_markup=main_menu(user_id))


@router.message(CommandStart())
@router.message(Command("menu"))
async def command_menu(message: Message, state: FSMContext) -> None:
    await cleanup_pending_excel(state)
    await state.clear()
    payload = (message.text or "").split(maxsplit=1)
    if len(payload) > 1 and payload[1].startswith("wait_"):
        try:
            wait_id = int(payload[1].removeprefix("wait_"))
        except ValueError:
            wait_id = 0
        entry = materials_db.get_wait_entry(wait_id)
        if entry and message.from_user and (
            entry.manager_id == message.from_user.id or is_admin(message.from_user.id)
        ):
            await message.answer(wait_entry_text(entry), reply_markup=wait_entry_keyboard(entry))
            return
    await show_main(message, message.from_user.id if message.from_user else None)


@router.message(Command("cancel"))
async def cancel(message: Message, state: FSMContext) -> None:
    await cleanup_pending_excel(state)
    await state.clear()
    await message.answer("✅ Текущее действие отменено.")
    await show_main(message, message.from_user.id if message.from_user else None)


@router.message(F.text.regexp(re.compile(r"^/прайс(?:@\w+)?\s*$", re.IGNORECASE)))
async def send_all_current_prices(message: Message, state: FSMContext) -> None:
    """Send the latest base price file for every warehouse."""
    await state.clear()
    warehouse_cities = {
        "center": "Москва",
        "ural": "Челябинск",
        "west": "Санкт-Петербург",
    }
    warehouse_order = ("center", "ural", "west")
    available = [
        warehouse
        for warehouse in warehouse_order
        if (price_storage_path / f"{warehouse}.xlsx").is_file()
    ]

    if not available:
        await message.answer(
            "📄 <b>Актуальные прайсы</b>\n\n"
            "Прайсы пока не загружены. Обратитесь к администратору."
        )
        return

    await message.answer(materials_db.get_setting(PRICE_MESSAGE_SETTING) or DEFAULT_PRICE_COMMAND_MESSAGE)
    for warehouse in available:
        warehouse_name = WAREHOUSES[warehouse]
        city = warehouse_cities[warehouse]
        source = price_storage_path / f"{warehouse}.xlsx"
        await message.answer_document(
            FSInputFile(source, filename=f"прайс {warehouse_name} — {city}.xlsx"),
            caption=f"📄 <b>{warehouse_name} — {city}</b>\nАктуальный базовый прайс.",
        )

    missing = [
        f"{WAREHOUSES[warehouse]} — {warehouse_cities[warehouse]}"
        for warehouse in warehouse_order
        if warehouse not in available
    ]
    if missing:
        await message.answer(
            "⚠️ Пока недоступны:\n" + "\n".join(f"• {html.escape(value)}" for value in missing)
        )


@router.message(Command("wait"))
async def add_wait_from_client_chat(message: Message, bot: Bot) -> None:
    if message.chat.type not in {"group", "supergroup"} or not message.from_user:
        await message.answer(
            "Команда <code>/wait</code> используется непосредственно в клиентском чате.\n\n"
            "Пример: <code>/wait картридж Vaporesso XROS 0.6 2мл</code>"
        )
        return
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or len(parts[1].strip()) < 2:
        await message.reply(
            "После команды укажите товар. Например:\n"
            "<code>/wait картридж Vaporesso XROS 0.6 2мл</code>"
        )
        return
    query, separator, explicit_comment = parts[1].partition("|")
    query = query.strip()
    comment = explicit_comment.strip() if separator else ""
    if not comment and message.reply_to_message:
        comment = (message.reply_to_message.text or message.reply_to_message.caption or "").strip()
    if len(query) < 2:
        await message.reply("После команды укажите название ожидаемого товара.")
        return
    if len(query) > 200:
        await message.reply("Описание слишком длинное. Оставьте бренд, модель и важные характеристики.")
        return
    title = clean_client_title(message.chat.title or str(message.chat.id))
    materials_db.upsert_client_chat(message.chat.id, title, message.chat.type, True)
    entry = materials_db.add_wait_entry(
        message.chat.id, title, message.from_user.id, message.from_user.full_name, query,
        message.message_id, comment[:500],
    )
    me = await bot.get_me()
    private_url = f"https://t.me/{me.username}?start=wait_{entry.id}"
    try:
        await bot.send_message(
            entry.manager_id,
            wait_entry_text(entry)
            + "\n\nПроверьте запрос и при необходимости добавьте комментарий.",
            reply_markup=wait_entry_keyboard(entry),
        )
        private_note = "Карточка отправлена менеджеру в личный диалог."
    except Exception:
        logging.exception("Не удалось открыть ожидание %s в личном диалоге", entry.id)
        private_note = "Нажмите кнопку ниже и запустите личный диалог с ботом."
    await message.reply(
        "🔔 <b>Запрос сохранён</b>\n\n" + private_note,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✏️ Дооформить в личном чате", url=private_url)]
        ]),
    )


@router.callback_query(F.data == "main:menu")
async def callback_menu(callback: CallbackQuery, state: FSMContext) -> None:
    await cleanup_pending_excel(state)
    await state.clear()
    await show_main(callback.message, callback.from_user.id, edit=True)
    await callback.answer()


@router.callback_query(F.data == "main:region")
async def open_region_search(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(AppState.region_search)
    await callback.message.edit_text(
        "🔎 <b>Поиск менеджера</b>\n\n"
        "Отправьте название города, посёлка, деревни, области или другой территории.\n\n"
        "Например: <code>Питер</code>",
        reply_markup=back_main(),
    )
    await callback.answer()


def unique_results(entries: list[Entry]) -> list[tuple[str, str, str]]:
    return list(dict.fromkeys((entry.name, entry.location, entry.manager) for entry in entries))


def manager_html(manager: str) -> str:
    escaped = html.escape(manager)
    username = MANAGER_LINKS.get(normalize(manager))
    if not username:
        return f"<b>{escaped}</b>"
    return f'<b><a href="https://t.me/{username}">{escaped}</a></b>'


def format_result(entries: list[Entry], page: int = 0) -> str:
    unique = unique_results(entries)
    if len(unique) == 1:
        name, location, manager = unique[0]
        lines = ["✅ <b>Менеджер найден</b>", "", f"📍 Территория: <b>{html.escape(name)}</b>"]
        if location:
            lines.append(f"🗺 Местоположение: <b>{html.escape(location)}</b>")
        lines.append(f"👤 Менеджер: {manager_html(manager)}")
        return "\n".join(lines)
    page_count = (len(unique) + RESULTS_PER_PAGE - 1) // RESULTS_PER_PAGE
    page = max(0, min(page, page_count - 1))
    start = page * RESULTS_PER_PAGE
    lines = ["🔎 <b>Найдено несколько вариантов</b>", f"Всего совпадений: <b>{len(unique)}</b>"]
    if page_count > 1:
        lines.append(f"Страница <b>{page + 1}</b> из <b>{page_count}</b>")
    for number, (name, location, manager) in enumerate(
        unique[start : start + RESULTS_PER_PAGE], start=start + 1
    ):
        block = [f"\n<b>{number}.</b> 📍 Территория: <b>{html.escape(name)}</b>"]
        if location:
            block.append(f"🗺 Местоположение: <b>{html.escape(location)}</b>")
        block.append(f"👤 Менеджер: {manager_html(manager)}")
        lines.append("\n".join(block))
    return "\n".join(lines)


def suggestion_label(entry: Entry, variant_count: int = 1) -> str:
    label = (
        f"{entry.name} — {variant_count} вариантов"
        if variant_count > 1
        else (f"{entry.name} — {entry.location}" if entry.location else entry.name)
    )
    return label if len(label) <= 64 else f"{label[:61]}..."


def pagination_keyboard(entry_index: int, entries: list[Entry], page: int) -> InlineKeyboardMarkup:
    total = len(unique_results(entries))
    page_count = (total + RESULTS_PER_PAGE - 1) // RESULTS_PER_PAGE
    back = f"page:{entry_index}:{page - 1}" if page > 0 else None
    forward = f"page:{entry_index}:{page + 1}" if page + 1 < page_count else None
    rows = [compact_nav(back, forward_data=forward)]
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(AppState.region_search, F.text)
async def search_region(message: Message) -> None:
    query = message.text.strip()
    if len(query) > 200:
        await message.answer(
            "⚠️ Запрос получился слишком длинным. Отправьте только название территории.",
            reply_markup=back_main(),
        )
        return
    found = catalog.exact(query)
    if found:
        index = catalog.by_name[normalize(query)][0]
        await message.answer(format_result(found), reply_markup=pagination_keyboard(index, found, 0))
        return
    indexes = catalog.suggestions(query)
    if not indexes:
        await message.answer(
            "🤷 <b>Ничего не найдено</b>\n\n"
            "Проверьте написание или попробуйте указать другое название территории.",
            reply_markup=back_main(),
        )
        return
    rows = []
    for index in indexes:
        entry = catalog.entries[index]
        count = len(catalog.by_name[normalize(entry.name)])
        rows.append([InlineKeyboardButton(text=suggestion_label(entry, count), callback_data=f"place:{index}")])
    rows.append(compact_nav())
    await message.answer(
        "Точного совпадения не найдено. Возможно, вы имели в виду один из вариантов:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.message(AppState.region_search)
async def region_non_text(message: Message) -> None:
    await message.answer(
        "Пожалуйста, отправьте название территории обычным текстовым сообщением.",
        reply_markup=back_main(),
    )


@router.callback_query(F.data.startswith("place:"))
async def select_suggestion(callback: CallbackQuery) -> None:
    try:
        index = int(callback.data.split(":", 1)[1])
        selected = catalog.entries[index]
    except (ValueError, IndexError):
        await callback.answer("Выполните поиск ещё раз.", show_alert=True)
        return
    found = catalog.exact(selected.name)
    await callback.message.edit_text(format_result(found), reply_markup=pagination_keyboard(index, found, 0))
    await callback.answer()


@router.callback_query(F.data.startswith("page:"))
async def change_page(callback: CallbackQuery) -> None:
    try:
        _, raw_index, raw_page = callback.data.split(":", 2)
        index, page = int(raw_index), int(raw_page)
        selected = catalog.entries[index]
    except (ValueError, IndexError):
        await callback.answer("Выполните поиск ещё раз.", show_alert=True)
        return
    found = catalog.exact(selected.name)
    await callback.message.edit_text(
        format_result(found, page), reply_markup=pagination_keyboard(index, found, page)
    )
    await callback.answer()


def money(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{quantized:,.2f}".replace(",", " ").replace(".", ",")


def discounted(value: Decimal, percent: int) -> Decimal:
    return value * (Decimal(100 - percent) / Decimal(100))


def variant_word(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return "вариант"
    if count % 10 in (2, 3, 4) and count % 100 not in (12, 13, 14):
        return "варианта"
    return "вариантов"


def price_group_label(display_name: str, category_name: str) -> str:
    label = f"{display_name} · {category_name}"
    return label if len(label) <= 64 else f"{label[:61]}..."


def price_search_keyboard(group_ids: list[int], page: int) -> InlineKeyboardMarkup:
    summaries = {group.callback_id: group for group in prices_db.group_summaries()}
    available = [summaries[group_id] for group_id in group_ids if group_id in summaries]
    page_count = max(1, (len(available) + PRICE_GROUPS_PER_PAGE - 1) // PRICE_GROUPS_PER_PAGE)
    page = max(0, min(page, page_count - 1))
    start = page * PRICE_GROUPS_PER_PAGE
    rows = [
        [
            InlineKeyboardButton(
                text=price_group_label(group.display_name, group.category_name),
                callback_data=f"price:g:{group.callback_id}",
            )
        ]
        for group in available[start : start + PRICE_GROUPS_PER_PAGE]
    ]
    back = f"price:r:{page - 1}" if page > 0 else None
    forward = f"price:r:{page + 1}" if page + 1 < page_count else None
    if back or forward:
        rows.append(compact_nav(back, forward_data=forward, home=False))
    rows.extend([
        [
            InlineKeyboardButton(text="➕ Добавить все", callback_data="price:add_all"),
            InlineKeyboardButton(text="🧺 Подборка", callback_data="price:cart"),
        ],
        compact_nav(search_data="main:prices"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def price_search_text(total: int, page: int) -> str:
    page_count = max(1, (total + PRICE_GROUPS_PER_PAGE - 1) // PRICE_GROUPS_PER_PAGE)
    page = max(0, min(page, page_count - 1))
    text = f"🔎 <b>Подходящих групп: {total}</b>"
    if page_count > 1:
        text += f"\nСтраница <b>{page + 1}</b> из <b>{page_count}</b>"
    return f"{text}\n\nВыберите нужную:"


def price_item_label(name: str) -> str:
    return name if len(name) <= 64 else f"{name[:61]}..."


def price_item_search_keyboard(item_ids: list[int], page: int) -> InlineKeyboardMarkup:
    summaries = {item.callback_id: item for item in prices_db.item_summaries()}
    available = [summaries[item_id] for item_id in item_ids if item_id in summaries]
    page_count = max(1, (len(available) + PRICE_ITEMS_PER_PAGE - 1) // PRICE_ITEMS_PER_PAGE)
    page = max(0, min(page, page_count - 1))
    start = page * PRICE_ITEMS_PER_PAGE
    rows = [[InlineKeyboardButton(
        text=price_item_label(item.name), callback_data=f"price:i:{item.callback_id}"
    )] for item in available[start : start + PRICE_ITEMS_PER_PAGE]]
    back = f"price:ir:{page - 1}" if page > 0 else None
    forward = f"price:ir:{page + 1}" if page + 1 < page_count else None
    if back or forward:
        rows.append(compact_nav(back, forward_data=forward, home=False))
    rows.append(compact_nav(search_data="main:prices"))
    return InlineKeyboardMarkup(inline_keyboard=rows)


def price_item_search_text(total: int, page: int) -> str:
    page_count = max(1, (total + PRICE_ITEMS_PER_PAGE - 1) // PRICE_ITEMS_PER_PAGE)
    page = max(0, min(page, page_count - 1))
    text = f"🔎 <b>Подходящих позиций: {total}</b>"
    if page_count > 1:
        text += f"\nСтраница <b>{page + 1}</b> из <b>{page_count}</b>"
    return f"{text}\n\nВыберите товар, чтобы увидеть цены и наличие:"


def format_price_group(details: GroupDetails) -> str:
    lines = [
        f"💰 <b>{html.escape(details.summary.display_name)}</b>",
        f"Категория: <b>{html.escape(details.summary.category_name)}</b>",
        f"Уникальных вариантов: <b>{details.unique_variants}</b>",
        "",
        "🏢 <b>Наличие и ассортимент</b>",
    ]
    for warehouse in ("center", "west", "ural"):
        count = details.summary.warehouse_counts.get(warehouse, 0)
        if count:
            lines.append(f"✅ {WAREHOUSES[warehouse]} — <b>{count}</b> {variant_word(count)}")
        else:
            lines.append(f"❌ {WAREHOUSES[warehouse]} — нет в прайсе")
    lines.extend(["", "💳 <b>Цены внутри группы</b>"])
    if len(details.tiers) > 1:
        cash_values = [tier.cash for tier in details.tiers]
        cashless_values = [tier.cashless for tier in details.tiers]
        lines.extend([
            "В этой группе цены зависят от конкретной позиции.",
            f"• Нал — <b>от {money(min(cash_values))} до {money(max(cash_values))} ₽</b>",
            f"• Безнал — <b>от {money(min(cashless_values))} до {money(max(cashless_values))} ₽</b>",
            "",
            "Откройте <b>«Позиции и цены»</b>, чтобы выбрать нужную модель или характеристику.",
        ])
        return "\n".join(lines)
    for number, tier in enumerate(details.tiers, 1):
        if len(details.tiers) > 1:
            lines.extend([
                "",
                f"<b>Ценовой уровень {number}</b> · {tier.variant_count} {variant_word(tier.variant_count)}",
            ])
        for percent in (0, 5, 10, 15):
            cash = money(discounted(tier.cash, percent))
            cashless = money(discounted(tier.cashless, percent))
            label = "Без скидки" if percent == 0 else f"Скидка {percent}%"
            lines.extend([
                "",
                f"<b>{label}</b>",
                f"• Нал — <b>{cash} ₽</b>",
                f"• Безнал — <b>{cashless} ₽</b>",
            ])
    return "\n".join(lines)


def format_price_variants(summary, variants, page: int) -> str:
    page_count = max(1, (len(variants) + PRICE_VARIANTS_PER_PAGE - 1) // PRICE_VARIANTS_PER_PAGE)
    page = max(0, min(page, page_count - 1))
    start = page * PRICE_VARIANTS_PER_PAGE
    shown = variants[start : start + PRICE_VARIANTS_PER_PAGE]
    lines = [
        f"📋 <b>{html.escape(summary.display_name)}</b>",
        f"Позиций: <b>{len(variants)}</b>",
    ]
    if page_count > 1:
        lines.append(f"Страница <b>{page + 1}</b> из <b>{page_count}</b>")
    for number, variant in enumerate(shown, start=start + 1):
        lines.extend([
            "",
            f"<b>{number}.</b> {html.escape(variant.name)}",
        ])
        unique_prices = set(variant.warehouse_prices.values())
        if len(unique_prices) == 1:
            warehouses = " · ".join(
                WAREHOUSES[key] for key in ("center", "west", "ural") if key in variant.warehouses
            )
            cash, cashless = next(iter(unique_prices))
            lines.extend([
                f"📍 {html.escape(warehouses)}",
                f"Нал <b>{money(cash)} ₽</b> · Безнал <b>{money(cashless)} ₽</b>",
            ])
        else:
            for warehouse in ("center", "west", "ural"):
                if warehouse in variant.warehouse_prices:
                    cash, cashless = variant.warehouse_prices[warehouse]
                    lines.append(
                        f"📍 {WAREHOUSES[warehouse]}: нал <b>{money(cash)} ₽</b> · "
                        f"безнал <b>{money(cashless)} ₽</b>"
                    )
    return "\n".join(lines)


def format_price_item(item) -> str:
    lines = [
        f"💰 <b>{html.escape(item.name)}</b>",
        f"Группа: <b>{html.escape(item.group_name)}</b>",
        f"Категория: <b>{html.escape(item.category_name)}</b>",
        "",
        "🏢 <b>Цены и наличие</b>",
    ]
    price_warehouses = {}
    for warehouse, prices in item.warehouse_prices.items():
        price_warehouses.setdefault(prices, []).append(warehouse)
    for (cash, cashless), warehouses in price_warehouses.items():
        names = " · ".join(WAREHOUSES[key] for key in ("center", "west", "ural") if key in warehouses)
        lines.extend(["", f"📍 <b>{html.escape(names)}</b>"])
        for percent in (0, 5, 10, 15):
            label = "Базовая" if percent == 0 else f"−{percent}%"
            lines.append(
                f"{label}: нал <b>{money(discounted(cash, percent))} ₽</b> · "
                f"безнал <b>{money(discounted(cashless, percent))} ₽</b>"
            )
    missing = [WAREHOUSES[key] for key in ("center", "west", "ural") if key not in item.warehouse_prices]
    if missing:
        lines.extend(["", f"❌ Нет в прайсе: {html.escape(' · '.join(missing))}"])
    return "\n".join(lines)


def wait_match_score(query: str, group_name: str, item_name: str) -> float:
    """Match a manager's human description without relying on a product code."""
    query_tokens = normalize_price_text(query).split()
    candidate = normalize_price_text(f"{group_name} {item_name}")
    candidate_tokens = candidate.split()
    if not query_tokens or not candidate_tokens:
        return 0.0
    # Категория — обязательное условие. Это не даёт запросу «картриджи XROS»
    # совпасть с устройствами XROS, а жидкостям OGGO — с ароматизаторами OGGO.
    category_rules = (
        (("жидкост",), ("жидкост",)),
        (("картридж",), ("картридж",)),
        (("ароматизатор", "аромамикс", "конструктор"), ("ароматизатор", "конструктор")),
        (("однораз",), ("однораз",)),
    )
    for query_markers, candidate_markers in category_rules:
        if any(token.startswith(query_markers) for token in query_tokens):
            if not any(token.startswith(candidate_markers) for token in candidate_tokens):
                return 0.0
            break
    scores = []
    for token in query_tokens:
        if token in candidate_tokens:
            score = 1.0
        elif token in candidate:
            score = 0.94
        elif any(char.isdigit() for char in token):
            # Модель, объём и сопротивление должны совпадать точно.
            return 0.0
        else:
            score = max(SequenceMatcher(None, token, word).ratio() for word in candidate_tokens)
        threshold = 0.82 if len(token) <= 2 else 0.68
        if score < threshold:
            return 0.0
        scores.append(score)
    return sum(scores) / len(scores)


async def notify_waitlist_matches(bot: Bot, reports: dict[str, dict] | None) -> int:
    """Notify managers immediately for the warehouse price that has just been applied."""
    if not reports:
        return 0
    arrivals = []
    for warehouse, report in reports.items():
        if warehouse not in WAREHOUSES:
            continue
        for item in report.get("added", []):
            arrivals.append((warehouse, item))
    if not arrivals:
        return 0
    sent = 0
    for entry in materials_db.list_wait_entries(active_only=True):
        matches = {}
        for warehouse, item in arrivals:
            search_group = f"{item.get('category', '')} {item['group']}"
            score = wait_match_score(entry.query, search_group, item["name"])
            if not score:
                continue
            item_key = normalize_price_text(item["name"])
            warehouse_signature = f"{warehouse}|{item_key}"
            if materials_db.wait_match_seen(entry.id, warehouse_signature):
                continue
            match = matches.setdefault(item_key, {
                "name": item["name"], "group": item["group"],
                "category": item.get("category", ""), "score": score,
                "warehouses": {}, "signatures": set(),
            })
            match["score"] = max(match["score"], score)
            match["warehouses"][warehouse] = (Decimal(item["cash"]), Decimal(item["cashless"]))
            match["signatures"].add(warehouse_signature)
        if not matches:
            continue
        group_wait = not any(char.isdigit() for char in normalize_price_text(entry.query))
        selected = sorted(matches.items(), key=lambda value: (-value[1]["score"], value[1]["name"]))
        if not group_wait:
            selected = selected[:1]
        warehouse_counts = {}
        payload_items = []
        for _, item in selected:
            warehouses = [WAREHOUSES[key] for key in ("center", "west", "ural") if key in item["warehouses"]]
            payload_items.append({"name": item["name"], "warehouses": warehouses})
            for warehouse in warehouses:
                warehouse_counts[warehouse] = warehouse_counts.get(warehouse, 0) + 1
        payload = {
            "count": len(selected),
            "warehouses": list(warehouse_counts),
            "warehouse_counts": warehouse_counts,
            "items": payload_items,
            "categories": sorted({item.get("category", "") for _, item in selected if item.get("category")}),
        }
        materials_db.set_wait_last_match(entry.id, payload)
        updated_entry = materials_db.get_wait_entry(entry.id)
        buttons = []
        chat_url = wait_chat_url(updated_entry)
        if chat_url:
            buttons.append([InlineKeyboardButton(text="➡️ Перейти в чат клиента", url=chat_url)])
        buttons.extend([
            [
                InlineKeyboardButton(text="📋 Подробнее", callback_data=f"wait:arrival:{entry.id}"),
                InlineKeyboardButton(text="✉️ Сообщение", callback_data=f"wait:message:{entry.id}"),
            ],
            [
                InlineKeyboardButton(text="✅ Сообщили", callback_data=f"wait:done:{entry.id}"),
                InlineKeyboardButton(text="⏳ Оставить", callback_data=f"wait:keep:{entry.id}"),
            ],
        ])
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        heading = (
            "🔔 <b>Поступила ожидаемая группа товаров</b>"
            if group_wait else "🔔 <b>Возможно, приехал ожидаемый товар</b>"
        )
        lines = [
            heading, "", f"Клиент: <b>{html.escape(clean_client_title(entry.client_title))}</b>",
            f"Ожидали: <b>{html.escape(entry.query)}</b>",
        ]
        if entry.comment:
            lines.append(f"Комментарий: {html.escape(entry.comment)}")
        lines.extend(["", f"Подходящих новых позиций: <b>{len(selected)}</b>"])
        for warehouse, count in warehouse_counts.items():
            lines.append(f"• {warehouse}: <b>{count}</b>")
        lines.extend(["", "Откройте подробности или подготовьте сообщение клиенту."])
        try:
            await bot.send_message(
                entry.manager_id,
                "\n".join(lines),
                reply_markup=keyboard,
            )
        except Exception:
            logging.exception("Не удалось уведомить менеджера %s об ожидании %s", entry.manager_id, entry.id)
            continue
        for _, item in selected:
            for signature in item["signatures"]:
                materials_db.record_wait_match(entry.id, signature)
        sent += 1
    return sent


def waitlist_keyboard(user_id: int) -> InlineKeyboardMarkup:
    entries = materials_db.list_wait_entries(manager_id=user_id, active_only=True)
    rows = [[InlineKeyboardButton(text="➕ Добавить ожидание", callback_data="wait:new")]]
    rows.extend([[InlineKeyboardButton(
        text=f"🔔 {clean_client_title(entry.client_title)[:22]} · {entry.query[:28]}",
        callback_data=f"wait:open:{entry.id}",
    )] for entry in entries[:20]])
    rows.append(compact_nav())
    return InlineKeyboardMarkup(inline_keyboard=rows)


def waitlist_text(user_id: int) -> str:
    entries = materials_db.list_wait_entries(manager_id=user_id, active_only=True)
    if not entries:
        return (
            "🔔 <b>Лист ожидания пока пуст</b>\n\n"
            "Чтобы зафиксировать запрос, напишите непосредственно в клиентском чате:\n\n"
            "<code>/wait картридж Vaporesso XROS 0.6 2мл</code>\n\n"
            "Либо нажмите <b>«Добавить ожидание»</b>, если клиентского чата нет."
        )
    shown = entries[:20]
    blocks = ["🔔 <b>Мой лист ожидания</b>"]
    for number, entry in enumerate(shown, 1):
        query = f"{entry.query[:100]}{'…' if len(entry.query) > 100 else ''}"
        comment = f"\n💬 {html.escape(entry.comment[:100])}" if entry.comment else ""
        blocks.append(
            f"<b>{number}. {html.escape(clean_client_title(entry.client_title))}</b>\n"
            f"{html.escape(query)}{comment}"
        )
    if len(entries) > len(shown):
        blocks.append(f"Показаны последние {len(shown)} из {len(entries)} запросов.")
    blocks.append(
        "Чтобы добавить запрос, напишите в нужном клиентском чате:\n"
        "<code>/wait название товара</code>\n\n"
        "Комментарий можно добавить после символа <code>|</code> или ответить командой на сообщение клиента."
    )
    return "\n\n".join(blocks)


def wait_chat_url(entry) -> str | None:
    chat_id = str(entry.chat_id)
    if chat_id.startswith("-100") and entry.source_message_id:
        return f"https://t.me/c/{chat_id[4:]}/{entry.source_message_id}"
    return None


def wait_entry_text(entry) -> str:
    mode = "Группа товаров" if not any(char.isdigit() for char in normalize_price_text(entry.query)) else "Конкретная позиция"
    lines = [
        "🔔 <b>Ожидание</b>", "",
        f"Клиент: <b>{html.escape(clean_client_title(entry.client_title))}</b>",
        f"Что ожидает: <b>{html.escape(entry.query)}</b>",
        f"Режим: <b>{mode}</b>",
        f"Комментарий: {html.escape(entry.comment) if entry.comment else '<i>не добавлен</i>'}",
    ]
    if entry.last_match:
        lines.extend([
            "", "✅ <b>Последнее найденное поступление</b>",
            f"Подходящих позиций: <b>{entry.last_match.get('count', 0)}</b>",
            f"Склады: {html.escape(' · '.join(entry.last_match.get('warehouses', [])))}",
        ])
    return "\n".join(lines)


def wait_entry_keyboard(entry) -> InlineKeyboardMarkup:
    rows = []
    chat_url = wait_chat_url(entry)
    if chat_url:
        rows.append([InlineKeyboardButton(text="➡️ Перейти в чат клиента", url=chat_url)])
    if entry.last_match:
        rows.append([InlineKeyboardButton(
            text="✉️ Сформировать сообщение клиенту", callback_data=f"wait:message:{entry.id}"
        )])
    rows.extend([
        [
            InlineKeyboardButton(text="✏️ Товар", callback_data=f"wait:edit_query:{entry.id}"),
            InlineKeyboardButton(text="💬 Комментарий", callback_data=f"wait:edit_comment:{entry.id}"),
        ],
        [InlineKeyboardButton(text="✅ Закрыть ожидание", callback_data=f"wait:done:{entry.id}")],
        compact_nav("main:waitlist"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "main:waitlist")
async def open_waitlist(callback: CallbackQuery) -> None:
    await callback.message.edit_text(
        waitlist_text(callback.from_user.id), reply_markup=waitlist_keyboard(callback.from_user.id)
    )
    await callback.answer()


@router.callback_query(F.data == "wait:new")
async def start_manual_wait(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(WaitState.new_client)
    await callback.message.edit_text(
        "➕ <b>Новое ожидание</b>\n\n"
        "Введите название клиента или компании.\n\n"
        "Например: <code>Vape Shop 24</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav("main:waitlist")]),
    )
    await callback.answer()


@router.message(WaitState.new_client, F.text)
async def save_manual_wait_client(message: Message, state: FSMContext) -> None:
    client = clean_client_title(message.text[:150])
    await state.update_data(wait_new_client=client)
    await state.set_state(WaitState.new_query)
    await message.answer(
        f"Клиент: <b>{html.escape(client)}</b>\n\n"
        "Что ожидает клиент?\n\n"
        "Например: <code>жидкости OGGO</code> или <code>картридж XROS 0.6 2мл</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav("main:waitlist")]),
    )


@router.message(WaitState.new_query, F.text)
async def save_manual_wait_query(message: Message, state: FSMContext) -> None:
    query = message.text.strip()[:200]
    if len(query) < 2:
        await message.answer("Укажите более понятное название товара.")
        return
    await state.update_data(wait_new_query=query)
    await state.set_state(WaitState.new_comment)
    await message.answer(
        "💬 Добавьте комментарий, если он нужен.\n\n"
        "Например: <code>клиенту нужно 20 коробок</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Без комментария", callback_data="wait:new_finish")],
            compact_nav("main:waitlist"),
        ]),
    )


async def finish_manual_wait(message: Message, state: FSMContext, manager, comment: str = "") -> None:
    data = await state.get_data()
    entry = materials_db.add_wait_entry(
        0, data["wait_new_client"], manager.id, manager.full_name,
        data["wait_new_query"], comment=comment[:500],
    )
    await state.clear()
    await message.answer("✅ <b>Ожидание создано</b>")
    await message.answer(wait_entry_text(entry), reply_markup=wait_entry_keyboard(entry))


@router.message(WaitState.new_comment, F.text)
async def save_manual_wait_comment(message: Message, state: FSMContext) -> None:
    await finish_manual_wait(message, state, message.from_user, message.text)


@router.callback_query(F.data == "wait:new_finish")
async def finish_manual_wait_without_comment(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await finish_manual_wait(callback.message, state, callback.from_user)


@router.callback_query(F.data.startswith("wait:open:"))
async def open_wait_entry(callback: CallbackQuery) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or (entry.manager_id != callback.from_user.id and not is_admin(callback.from_user.id)):
        await callback.answer("Ожидание недоступно.", show_alert=True)
        return
    await callback.message.edit_text(wait_entry_text(entry), reply_markup=wait_entry_keyboard(entry))
    await callback.answer()


@router.callback_query(F.data.startswith("wait:edit_query:"))
async def start_wait_query_edit(callback: CallbackQuery, state: FSMContext) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or entry.manager_id != callback.from_user.id:
        await callback.answer("Ожидание недоступно.", show_alert=True)
        return
    await state.set_state(WaitState.edit_query)
    await state.update_data(wait_edit_id=wait_id)
    await callback.message.edit_text(
        "✏️ <b>Что ожидает клиент?</b>\n\n"
        f"Сейчас: <b>{html.escape(entry.query)}</b>\n\n"
        "Отправьте новое описание. Например: <code>картриджи XROS</code>.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav(f"wait:open:{wait_id}")]),
    )
    await callback.answer()


@router.message(WaitState.edit_query, F.text)
async def save_wait_query_edit(message: Message, state: FSMContext) -> None:
    wait_id = (await state.get_data()).get("wait_edit_id")
    entry = materials_db.update_wait_entry(wait_id, message.from_user.id, query=message.text[:200])
    await state.clear()
    if not entry:
        await message.answer("Ожидание уже недоступно.", reply_markup=back_main())
        return
    await message.answer("✅ Описание обновлено.")
    await message.answer(wait_entry_text(entry), reply_markup=wait_entry_keyboard(entry))


@router.callback_query(F.data.startswith("wait:edit_comment:"))
async def start_wait_comment_edit(callback: CallbackQuery, state: FSMContext) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or entry.manager_id != callback.from_user.id:
        await callback.answer("Ожидание недоступно.", show_alert=True)
        return
    await state.set_state(WaitState.edit_comment)
    await state.update_data(wait_edit_id=wait_id)
    await callback.message.edit_text(
        "💬 <b>Комментарий к ожиданию</b>\n\n"
        "Отправьте новый комментарий. Чтобы удалить его, отправьте <code>-</code>.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav(f"wait:open:{wait_id}")]),
    )
    await callback.answer()


@router.message(WaitState.edit_comment, F.text)
async def save_wait_comment_edit(message: Message, state: FSMContext) -> None:
    wait_id = (await state.get_data()).get("wait_edit_id")
    comment = "" if message.text.strip() == "-" else message.text[:500]
    entry = materials_db.update_wait_entry(wait_id, message.from_user.id, comment=comment)
    await state.clear()
    if not entry:
        await message.answer("Ожидание уже недоступно.", reply_markup=back_main())
        return
    await message.answer("✅ Комментарий обновлён.")
    await message.answer(wait_entry_text(entry), reply_markup=wait_entry_keyboard(entry))


@router.callback_query(F.data.startswith("wait:arrival:"))
async def show_wait_arrival(callback: CallbackQuery) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or not entry.last_match or (
        entry.manager_id != callback.from_user.id and not is_admin(callback.from_user.id)
    ):
        await callback.answer("Подробности поступления недоступны.", show_alert=True)
        return
    lines = [
        "📋 <b>Поступление по ожиданию</b>", "",
        f"Клиент: <b>{html.escape(clean_client_title(entry.client_title))}</b>",
        f"Запрос: <b>{html.escape(entry.query)}</b>", "",
    ]
    items = entry.last_match.get("items", [])
    for number, item in enumerate(items[:25], 1):
        lines.append(
            f"<b>{number}.</b> {html.escape(item['name'])}\n"
            f"📍 {html.escape(' · '.join(item.get('warehouses', [])))}"
        )
    if len(items) > 25:
        lines.extend(["", f"И ещё позиций: <b>{len(items) - 25}</b>"])
    await callback.message.edit_text("\n\n".join(lines), reply_markup=wait_entry_keyboard(entry))
    await callback.answer()


def client_wait_message(entry) -> str:
    match = entry.last_match or {}
    group_wait = not any(char.isdigit() for char in normalize_price_text(entry.query))
    if group_wait:
        categories = match.get("categories", [])
        category = categories[0] if len(categories) == 1 else ""
        query = entry.query.strip()
        if category == "Жидкости":
            tail = re.sub(r"^жидкост(?:ь|и)\s*", "", query, flags=re.IGNORECASE).strip()
            subject = f"жидкости {tail}".strip()
            assortment = "разные линейки и вкусы"
        elif category == "Одноразовые системы":
            tail = re.sub(r"^(?:одноразовые\s+системы|одноразки)\s*", "", query, flags=re.IGNORECASE).strip()
            subject = f"одноразовые системы {tail}".strip()
            assortment = "разные модели и вкусы"
        elif category == "Электронные системы":
            tail = re.sub(r"^(?:устройства|электронные\s+системы)\s*", "", query, flags=re.IGNORECASE).strip()
            subject = f"устройства {tail}".strip()
            assortment = "разные модели и цвета"
        elif category in {"Картриджи", "Расходники"} and "картридж" in normalize_price_text(query):
            tail = re.sub(r"^картридж(?:и)?\s*", "", query, flags=re.IGNORECASE).strip()
            subject = f"картриджи {tail}".strip()
            assortment = "варианты с разным сопротивлением и объёмом"
        elif len(categories) > 1:
            subject = f"продукция {query}"
            assortment = "разные линейки и варианты"
        else:
            subject = query
            assortment = "разные варианты"
        text = (
            f"Добрый день! К нам поступили {subject} — в наличии {assortment}. "
            "Подскажите, товар ещё нужен? Если да, отправим цены и покажем, что сейчас есть в наличии."
        )
    else:
        item_name = entry.query[:1].upper() + entry.query[1:]
        text = (
            f"Добрый день! {item_name} снова в наличии. "
            "Подскажите, товар ещё нужен? Если да, отправим цены."
        )
    return text


@router.callback_query(F.data.startswith("wait:message:"))
async def prepare_client_wait_message(callback: CallbackQuery) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or not entry.last_match or (
        entry.manager_id != callback.from_user.id and not is_admin(callback.from_user.id)
    ):
        await callback.answer("Сначала должно быть найдено поступление.", show_alert=True)
        return
    text = client_wait_message(entry)
    rows = []
    chat_url = wait_chat_url(entry)
    if chat_url:
        rows.append([InlineKeyboardButton(text="➡️ Перейти в чат клиента", url=chat_url)])
    rows.append(compact_nav(f"wait:open:{entry.id}"))
    await callback.message.answer(
        "✉️ <b>Готовое сообщение клиенту</b>\n\n"
        "Нажмите на текст, чтобы скопировать его:\n\n"
        f"<pre>{html.escape(text)}</pre>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("wait:done:"))
async def close_wait_entry(callback: CallbackQuery) -> None:
    try:
        wait_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    manager_id = None if is_admin(callback.from_user.id) else callback.from_user.id
    if not materials_db.close_wait_entry(wait_id, manager_id):
        await callback.answer("Запрос уже закрыт или принадлежит другому менеджеру.", show_alert=True)
        return
    await callback.answer("Ожидание закрыто")
    if callback.message.text and "Мой лист ожидания" in callback.message.text:
        await callback.message.edit_text(
            waitlist_text(callback.from_user.id), reply_markup=waitlist_keyboard(callback.from_user.id)
        )
    else:
        await callback.message.edit_text(
            "✅ <b>Ожидание закрыто</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav("main:waitlist")]),
        )


@router.callback_query(F.data.startswith("wait:keep:"))
async def keep_wait_entry(callback: CallbackQuery) -> None:
    wait_id = int(callback.data.rsplit(":", 1)[1])
    entry = materials_db.get_wait_entry(wait_id)
    if not entry or (entry.manager_id != callback.from_user.id and not is_admin(callback.from_user.id)):
        await callback.answer("Этот запрос вам недоступен.", show_alert=True)
        return
    await callback.answer("Ожидание оставлено активным")


def price_variants_keyboard(callback_id: int, page: int, total: int) -> InlineKeyboardMarkup:
    page_count = max(1, (total + PRICE_VARIANTS_PER_PAGE - 1) // PRICE_VARIANTS_PER_PAGE)
    back = f"price:v:{callback_id}:{page - 1}" if page > 0 else None
    forward = f"price:v:{callback_id}:{page + 1}" if page + 1 < page_count else None
    rows = [compact_nav(back, forward_data=forward, home=False)] if back or forward else []
    rows.extend([
        [InlineKeyboardButton(text="💰 Вернуться к ценам", callback_data=f"price:g:{callback_id}")],
        compact_nav("price:back", search_data="main:prices"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def downloadable_prices_keyboard() -> InlineKeyboardMarkup:
    statuses = prices_db.import_statuses()
    buttons = [
            InlineKeyboardButton(
                text=f"🏢 {WAREHOUSES[warehouse]}",
                callback_data=f"files:w:{warehouse}",
            )
        for warehouse in ("center", "west", "ural")
        if warehouse in statuses
    ]
    rows = button_grid(buttons)
    rows.append(compact_nav())
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.my_chat_member()
async def track_client_chat(event: ChatMemberUpdated) -> None:
    if event.chat.type not in {"group", "supergroup"}:
        return
    active = event.new_chat_member.status in {"member", "administrator", "creator"}
    materials_db.upsert_client_chat(
        event.chat.id, clean_client_title(event.chat.title or str(event.chat.id)), event.chat.type, active
    )


def broadcast_menu_keyboard(user_id: int) -> InlineKeyboardMarkup:
    rows = [[
        InlineKeyboardButton(text="➕ Создать", callback_data="broadcast:new"),
        InlineKeyboardButton(text="📥 Список чатов", callback_data="broadcast:export_chats"),
    ], compact_nav()]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def service_chat_text() -> str:
    chat = materials_db.get_client_chat(SERVICE_CHAT_ID)
    return chat.title if chat else str(SERVICE_CHAT_ID)


@router.callback_query(F.data == "main:broadcasts")
async def open_broadcasts(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_broadcaster(callback): return
    await state.clear()
    await callback.message.edit_text(
        "📣 <b>Рассылки клиентам</b>\n\n"
        f"Зарегистрировано чатов: <b>{len(materials_db.list_client_chats())}</b>\n"
        f"Служебный чат: <b>{html.escape(service_chat_text())}</b>\n\n"
        "Аудитория каждой рассылки определяется новым Excel-файлом с колонкой Chat ID.",
        reply_markup=broadcast_menu_keyboard(callback.from_user.id),
    )
    await callback.answer()


def build_chats_excel(destination: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Чаты"
    sheet.append(["Название чата", "Chat ID", "Статус"])
    for chat in materials_db.list_client_chats():
        sheet.append([chat.title, chat.chat_id, "Активен" if chat.is_active else "Бот удалён"])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 55
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 18
    workbook.save(destination)
    workbook.close()


@router.callback_query(F.data == "broadcast:export_chats")
async def export_client_chats(callback: CallbackQuery) -> None:
    if not await require_broadcaster(callback): return
    await callback.answer("Готовлю список…")
    with tempfile.TemporaryDirectory() as temp_name:
        destination = Path(temp_name) / "список клиентских чатов.xlsx"
        await asyncio.to_thread(build_chats_excel, destination)
        await callback.message.answer_document(
            FSInputFile(destination, filename=destination.name),
            caption=f"📥 Чатов в реестре: <b>{len(materials_db.list_client_chats())}</b>",
        )


@router.callback_query(F.data == "broadcast:new")
async def new_broadcast(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_broadcaster(callback): return
    await state.clear()
    await state.set_state(BroadcastState.audience_upload)
    await callback.message.edit_text(
        "📄 <b>Шаг 1 из 3 · Получатели</b>\n\n"
        "Отправьте Excel-файл <code>.xlsx</code>. Бот найдёт колонку <b>Chat ID</b>, "
        "удалит дубли и подготовит список получателей.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отменить", callback_data="main:broadcasts")]
        ]),
    )
    await callback.answer()


def parse_audience_excel(path: Path) -> tuple[list[int], int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        workbook.close(); raise ValueError("Excel-файл пуст.")
    normalized = [normalize(str(value or "")) for value in header]
    column = next((i for i, value in enumerate(normalized) if value in {"chat id", "чат id", "id чата"}), None)
    if column is None:
        workbook.close(); raise ValueError("Не найдена колонка Chat ID.")
    values = []
    invalid = 0
    for row in rows:
        raw = row[column] if len(row) > column else None
        if raw is None or str(raw).strip() == "": continue
        try:
            chat_id = int(float(raw)) if isinstance(raw, float) else int(str(raw).strip())
            if chat_id >= 0: raise ValueError
            values.append(chat_id)
        except (ValueError, TypeError):
            invalid += 1
    workbook.close()
    unique = list(dict.fromkeys(values))
    return unique, len(values) - len(unique), invalid


@router.message(BroadcastState.audience_upload)
async def receive_broadcast_audience(message: Message, state: FSMContext, bot: Bot) -> None:
    if not can_broadcast(message.from_user.id, message.from_user.username): return
    if not message.document or not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer("⚠️ Отправьте документ в формате <code>.xlsx</code> с колонкой Chat ID."); return
    with tempfile.TemporaryDirectory() as temp_name:
        path = Path(temp_name) / "audience.xlsx"
        await bot.download(message.document.file_id, destination=path)
        try:
            chat_ids, duplicates, invalid = await asyncio.to_thread(parse_audience_excel, path)
        except Exception as error:
            await message.answer(f"❌ Не удалось прочитать список: {html.escape(str(error))}"); return
    if not chat_ids:
        await message.answer("❌ В файле нет корректных отрицательных Chat ID групп."); return
    registry = {chat.chat_id: chat for chat in materials_db.list_client_chats()}
    inactive = sum(chat_id in registry and not registry[chat_id].is_active for chat_id in chat_ids)
    unknown = sum(chat_id not in registry for chat_id in chat_ids)
    await state.update_data(broadcast_chat_ids=chat_ids)
    await state.set_state(BroadcastState.content_upload)
    await message.answer(
        "✅ <b>Список получателей подготовлен</b>\n\n"
        f"Уникальных Chat ID: <b>{len(chat_ids)}</b>\n"
        f"Удалено дублей: <b>{duplicates}</b>\n"
        f"Некорректных строк: <b>{invalid}</b>\n"
        f"Неизвестных боту чатов: <b>{unknown}</b>\n"
        f"Отмечены недоступными: <b>{inactive}</b>\n\n"
        "📨 <b>Шаг 2 из 3</b>\nОтправьте готовый пост: текст, фото, видео или документ с подписью.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отменить", callback_data="main:broadcasts")]
        ]),
    )


@router.message(BroadcastState.content_upload)
async def receive_broadcast_content(message: Message, state: FSMContext) -> None:
    if not can_broadcast(message.from_user.id, message.from_user.username): return
    if not (message.text or message.photo or message.video or message.document or message.animation):
        await message.answer("⚠️ Поддерживаются текст, фото, видео, анимация или документ."); return
    await state.update_data(source_chat_id=message.chat.id, source_message_id=message.message_id)
    await state.set_state(BroadcastState.ready)
    await message.answer("👁 <b>Предпросмотр поста:</b>")
    await message.copy_to(message.chat.id)
    await message.answer(
        "🧪 <b>Шаг 3 из 3</b>\n\nСначала отправьте тест в служебную группу.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="🧪 Тест", callback_data="broadcast:test"),
                InlineKeyboardButton(text="✏️ Заменить", callback_data="broadcast:replace_content"),
            ],
            [InlineKeyboardButton(text="❌ Отменить", callback_data="main:broadcasts")],
        ]),
    )


@router.callback_query(F.data == "broadcast:replace_content")
async def replace_broadcast_content(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_broadcaster(callback): return
    await state.set_state(BroadcastState.content_upload)
    await callback.message.edit_text("✏️ Отправьте новый вариант поста.")
    await callback.answer()


@router.callback_query(F.data == "broadcast:test")
async def test_broadcast(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    if not await require_broadcaster(callback): return
    data = await state.get_data()
    try:
        await bot.copy_message(SERVICE_CHAT_ID, data["source_chat_id"], data["source_message_id"])
    except Exception:
        logging.exception("Не удалось отправить тест рассылки")
        await callback.answer("Не удалось отправить тест. Проверьте доступ бота к служебной группе.", show_alert=True); return
    await callback.message.edit_text(
        "✅ <b>Тест отправлен в служебную группу</b>\n\n"
        f"Получателей в Excel: <b>{len(data.get('broadcast_chat_ids', []))}</b>\n\n"
        "Проверьте сообщение в группе и подтвердите массовую рассылку.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🚀 Начать рассылку", callback_data="broadcast:confirm_send")],
            [
                InlineKeyboardButton(text="✏️ Заменить", callback_data="broadcast:replace_content"),
                InlineKeyboardButton(text="❌ Отменить", callback_data="main:broadcasts"),
            ],
        ]),
    )
    await callback.answer()


def build_broadcast_report(destination: Path, results: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результат"
    sheet.append(["Название чата", "Chat ID", "Результат", "Ошибка"])
    for item in results:
        sheet.append([item["title"], item["chat_id"], item["status"], item["error"]])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in zip(("A", "B", "C", "D"), (55, 24, 18, 70)):
        sheet.column_dimensions[column].width = width
    workbook.save(destination)
    workbook.close()


@router.callback_query(F.data == "broadcast:confirm_send")
async def run_client_broadcast(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    if not await require_broadcaster(callback): return
    operator_id = callback.from_user.id
    if operator_id in broadcasts_in_progress:
        await callback.answer("Рассылка уже выполняется.", show_alert=True); return
    data = await state.get_data()
    chat_ids = data.get("broadcast_chat_ids", [])
    source_chat_id = data.get("source_chat_id")
    source_message_id = data.get("source_message_id")
    if not chat_ids or not source_chat_id or not source_message_id:
        await state.clear()
        await callback.answer("Черновик рассылки устарел. Создайте его заново.", show_alert=True); return
    broadcasts_in_progress.add(operator_id)
    await callback.answer()
    await callback.message.edit_text(
        "⏳ <b>Рассылка запущена</b>\n\n"
        f"Обработано: <b>0 из {len(chat_ids)}</b>\n"
        "Не запускайте вторую рассылку до завершения этой."
    )
    registry = {chat.chat_id: chat for chat in materials_db.list_client_chats()}
    results = []
    sent = failed = 0
    try:
        for number, chat_id in enumerate(chat_ids, 1):
            chat = registry.get(chat_id)
            title = chat.title if chat else "Неизвестный чат"
            if chat and not chat.is_active:
                results.append({"title": title, "chat_id": chat_id, "status": "Пропущен", "error": "Бот удалён из чата"})
                failed += 1
            else:
                try:
                    await bot.copy_message(chat_id, source_chat_id, source_message_id)
                    results.append({"title": title, "chat_id": chat_id, "status": "Отправлено", "error": ""})
                    sent += 1
                except TelegramRetryAfter as error:
                    await asyncio.sleep(error.retry_after)
                    try:
                        await bot.copy_message(chat_id, source_chat_id, source_message_id)
                        results.append({"title": title, "chat_id": chat_id, "status": "Отправлено", "error": ""})
                        sent += 1
                    except Exception as retry_error:
                        results.append({"title": title, "chat_id": chat_id, "status": "Ошибка", "error": str(retry_error)[:500]})
                        failed += 1
                except Exception as error:
                    results.append({"title": title, "chat_id": chat_id, "status": "Ошибка", "error": str(error)[:500]})
                    failed += 1
            if number % 20 == 0 and number < len(chat_ids):
                await edit_or_answer(
                    callback.message,
                    "⏳ <b>Рассылка выполняется</b>\n\n"
                    f"Обработано: <b>{number} из {len(chat_ids)}</b>\n"
                    f"Успешно: <b>{sent}</b> · Ошибок: <b>{failed}</b>"
                )
            await asyncio.sleep(0.05)
    finally:
        broadcasts_in_progress.discard(operator_id)
    await state.clear()
    await edit_or_answer(
        callback.message,
        "✅ <b>Рассылка завершена</b>\n\n"
        f"Всего чатов: <b>{len(chat_ids)}</b>\n"
        f"Успешно отправлено: <b>{sent}</b>\n"
        f"Ошибок и пропусков: <b>{failed}</b>\n\n"
        "Подробности находятся в Excel-отчёте.",
        reply_markup=broadcast_menu_keyboard(callback.from_user.id),
    )
    with tempfile.TemporaryDirectory() as temp_name:
        destination = Path(temp_name) / f"отчёт рассылки {datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        await asyncio.to_thread(build_broadcast_report, destination, results)
        await callback.message.answer_document(
            FSInputFile(destination, filename=destination.name), caption="📊 Подробный отчёт по рассылке"
        )


def combined_report_bundle() -> tuple[dict[str, dict] | None, str | None]:
    reports = prices_db.latest_reports()
    required = ("center", "west", "ural")
    if any(warehouse not in reports for warehouse in required):
        return None, "Сначала загрузите свежие прайсы всех трёх складов."
    today = datetime.now().astimezone().date().isoformat()
    if any(str(reports[warehouse].get("created_at", "")).split("T", 1)[0] != today for warehouse in required):
        return None, "Для общей рассылки нужны три отчёта, сформированные сегодня."
    return {warehouse: reports[warehouse] for warehouse in required}, None


def combined_report_signature(reports: dict[str, dict]) -> str:
    return "|".join(
        f"{warehouse}:{reports[warehouse].get('created_at')}:{reports[warehouse].get('current_file')}"
        for warehouse in ("center", "west", "ural")
    )


def price_reports_keyboard(user_id: int | None = None) -> InlineKeyboardMarkup:
    reports = prices_db.latest_reports()
    buttons = [InlineKeyboardButton(
            text=f"📊 {WAREHOUSES[warehouse]}", callback_data=f"reports:show:{warehouse}"
        )
        for warehouse in ("center", "west", "ural") if warehouse in reports
    ]
    rows = button_grid(buttons)
    combined, _ = combined_report_bundle()
    if is_admin(user_id) and combined:
        signature = combined_report_signature(combined)
        if not prices_db.report_broadcast_sent(signature):
            rows.append([InlineKeyboardButton(
                text="📣 Отправить общий отчёт", callback_data="reports:broadcast_all"
            )])
    rows.append(compact_nav())
    return InlineKeyboardMarkup(inline_keyboard=rows)


def format_price_report(report: dict) -> str:
    increased = sum(
        Decimal(item["new_cash"]) > Decimal(item["old_cash"])
        or Decimal(item["new_cashless"]) > Decimal(item["old_cashless"])
        for item in report["price_changes"]
    )
    decreased = sum(
        Decimal(item["new_cash"]) < Decimal(item["old_cash"])
        or Decimal(item["new_cashless"]) < Decimal(item["old_cashless"])
        for item in report["price_changes"]
    )
    previous_date = str(report["previous_date"]).split("T", 1)[0]
    current_date = str(report["current_date"]).split("T", 1)[0]
    return (
        f"📊 <b>Изменения прайса · {WAREHOUSES[report['warehouse']]}</b>\n\n"
        f"Сравнение: <b>{html.escape(previous_date)}</b> → <b>{html.escape(current_date)}</b>\n"
        f"Позиций: <b>{report['previous_count']}</b> → <b>{report['current_count']}</b>\n\n"
        f"➕ Появилось: <b>{len(report['added'])}</b>\n"
        f"❌ Закончилось: <b>{len(report['removed'])}</b>\n"
        f"📈 Подорожало: <b>{increased}</b>\n"
        f"📉 Подешевело: <b>{decreased}</b>\n"
        f"🆕 Новых групп: <b>{len(report['added_groups'])}</b>\n"
        f"⛔ Исчезнувших групп: <b>{len(report['removed_groups'])}</b>\n\n"
        "Подробный список товаров и изменения обеих цен находятся в Excel-отчёте."
    )


def report_actions_keyboard(warehouse: str) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(
        text="📥 Скачать подробный Excel", callback_data=f"reports:file:{warehouse}"
    )]]
    rows.append(compact_nav("main:price_reports"))
    return InlineKeyboardMarkup(inline_keyboard=rows)


def format_combined_report_notification(reports: dict[str, dict]) -> str:
    lines = ["🔔 <b>Прайсы обновлены</b>", "", "Свежие цены и остатки загружены по всем складам."]
    total_added = total_removed = total_prices = 0
    for warehouse in ("center", "west", "ural"):
        report = reports[warehouse]
        added, removed = len(report["added"]), len(report["removed"])
        changed = len(report["price_changes"])
        total_added += added
        total_removed += removed
        total_prices += changed
        lines.extend([
            "", f"🏢 <b>{WAREHOUSES[warehouse]}</b>",
            f"➕ Появилось: <b>{added}</b> · ❌ Закончилось: <b>{removed}</b>",
            f"💰 Изменилось цен: <b>{changed}</b>",
        ])
    lines.extend([
        "", "📊 <b>Итого по трём складам</b>",
        f"➕ Появилось: <b>{total_added}</b>",
        f"❌ Закончилось: <b>{total_removed}</b>",
        f"💰 Изменилось цен: <b>{total_prices}</b>",
        "", "Подробные отчёты доступны в разделе «Изменения прайсов».",
    ])
    return "\n".join(lines)


@router.callback_query(F.data == "main:price_reports")
async def open_price_reports(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    reports = prices_db.latest_reports()
    if not reports:
        await callback.message.edit_text(
            "📊 <b>Изменения прайсов</b>\n\n"
            "Отчётов пока нет. Первый отчёт появится после следующего обновления склада, "
            "для которого уже загружен предыдущий прайс.",
            reply_markup=back_main(),
        )
    else:
        await callback.message.edit_text(
            "📊 <b>Изменения прайсов</b>\n\n"
            "Здесь хранится только последнее сравнение для каждого склада. Выберите склад:",
            reply_markup=price_reports_keyboard(callback.from_user.id),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("reports:show:"))
async def show_price_report(callback: CallbackQuery) -> None:
    warehouse = callback.data.rsplit(":", 1)[1]
    report = prices_db.latest_report(warehouse)
    if warehouse not in WAREHOUSES or not report:
        await callback.answer("Отчёт этого склада пока недоступен.", show_alert=True)
        return
    await callback.message.edit_text(
        format_price_report(report),
        reply_markup=report_actions_keyboard(warehouse),
    )
    await callback.answer()


@router.callback_query(F.data == "reports:broadcast_all")
async def confirm_report_broadcast(callback: CallbackQuery) -> None:
    if not await require_admin(callback):
        return
    reports, error = combined_report_bundle()
    if not reports:
        await callback.answer(error or "Общий отчёт пока недоступен.", show_alert=True)
        return
    signature = combined_report_signature(reports)
    if prices_db.report_broadcast_sent(signature):
        await callback.answer("Этот общий отчёт уже был отправлен.", show_alert=True)
        return
    users = materials_db.list_access_users()
    ready = len({user.telegram_id for user in users if user.telegram_id})
    waiting = sum(user.telegram_id is None for user in users)
    await callback.message.edit_text(
        "📣 <b>Предпросмотр уведомления</b>\n\n"
        + format_combined_report_notification(reports)
        + f"\n\nПолучателей: <b>{ready}</b>"
        + (f"\nБез Telegram ID: <b>{waiting}</b> — будут пропущены" if waiting else "")
        + "\n\nОтправить это уведомление пользователям белого списка?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Отправить", callback_data="reports:send_all"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="main:price_reports"),
            ],
        ]),
    )
    await callback.answer()


@router.callback_query(F.data == "reports:send_all")
async def send_report_broadcast(callback: CallbackQuery, bot: Bot) -> None:
    if not await require_admin(callback):
        return
    reports, error = combined_report_bundle()
    if not reports:
        await callback.answer(error or "Общий отчёт пока недоступен.", show_alert=True)
        return
    signature = combined_report_signature(reports)
    if prices_db.report_broadcast_sent(signature):
        await callback.answer("Этот общий отчёт уже был отправлен.", show_alert=True)
        return
    recipients = sorted({
        user.telegram_id for user in materials_db.list_access_users() if user.telegram_id
    })
    missing_ids = sum(user.telegram_id is None for user in materials_db.list_access_users())
    await callback.answer("Начинаю рассылку…")
    await callback.message.edit_text(
        "⏳ <b>Отправляю уведомление</b>\n\n"
        "Отчёт: <b>Москва · Санкт-Петербург · Челябинск</b>\n"
        f"Получателей: <b>{len(recipients)}</b>"
    )
    sent = 0
    failed = 0
    markup = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Открыть отчёты", callback_data="main:price_reports")]
    ])
    for telegram_id in recipients:
        try:
            await bot.send_message(telegram_id, format_combined_report_notification(reports), reply_markup=markup)
            sent += 1
        except TelegramRetryAfter as error:
            await asyncio.sleep(error.retry_after)
            try:
                await bot.send_message(telegram_id, format_combined_report_notification(reports), reply_markup=markup)
                sent += 1
            except Exception:
                failed += 1
                logging.exception("Не удалось отправить отчёт пользователю %s", telegram_id)
        except Exception:
            failed += 1
            logging.exception("Не удалось отправить отчёт пользователю %s", telegram_id)
    if sent:
        prices_db.mark_report_broadcast_sent(signature)
    await callback.message.edit_text(
        "✅ <b>Рассылка завершена</b>\n\n"
        f"Успешно отправлено: <b>{sent}</b>\n"
        f"Не удалось отправить: <b>{failed}</b>\n"
        f"Без привязанного Telegram ID: <b>{missing_ids}</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            compact_nav("main:price_reports")
        ]),
    )


@router.callback_query(F.data.startswith("reports:file:"))
async def download_price_report(callback: CallbackQuery) -> None:
    warehouse = callback.data.rsplit(":", 1)[1]
    report = prices_db.latest_report(warehouse)
    if warehouse not in WAREHOUSES or not report:
        await callback.answer("Отчёт этого склада пока недоступен.", show_alert=True)
        return
    await callback.answer("Готовлю отчёт…")
    try:
        with tempfile.TemporaryDirectory() as temp_name:
            date = str(report["current_date"]).split("T", 1)[0]
            destination = Path(temp_name) / f"изменения прайса {WAREHOUSES[warehouse]} {date}.xlsx"
            await asyncio.to_thread(generate_change_report_excel, report, destination)
            await callback.message.answer_document(
                FSInputFile(destination, filename=destination.name),
                caption=f"📊 <b>Изменения прайса · {WAREHOUSES[warehouse]}</b>",
            )
    except Exception:
        logging.exception("Не удалось сформировать отчёт прайса %s", warehouse)
        await callback.message.answer("⚠️ Не удалось сформировать отчёт. Попробуйте ещё раз.")


@router.callback_query(F.data == "main:price_files")
async def open_downloadable_prices(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not prices_db.import_statuses():
        await callback.answer("Прайсы пока не загружены.", show_alert=True)
        return
    await callback.message.edit_text(
        "📄 <b>Прайсы</b>\n\n"
        "Выберите склад. Можно скачать исходный прайс с базовыми ценами или версию, "
        "в которой цены нал/безнал уже уменьшены на 10%.",
        reply_markup=downloadable_prices_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("files:w:"))
async def select_price_warehouse(callback: CallbackQuery) -> None:
    warehouse = callback.data.rsplit(":", 1)[1]
    status = prices_db.import_statuses().get(warehouse)
    if warehouse not in WAREHOUSES or not status:
        await callback.answer("Прайс этого склада пока недоступен.", show_alert=True)
        return
    price_date = (status["price_date"] or status["updated_at"] or "").split("T", 1)[0]
    await callback.message.edit_text(
        f"📄 <b>{WAREHOUSES[warehouse]}</b>\n\n"
        f"Дата прайса: <b>{html.escape(price_date)}</b>\n"
        f"Товарных позиций: <b>{status['item_count']}</b>\n\n"
        "Выберите нужный вариант:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="Базовые", callback_data=f"files:get:{warehouse}:base"),
                InlineKeyboardButton(text="Скидка −10%", callback_data=f"files:get:{warehouse}:10"),
            ],
            compact_nav("main:price_files"),
        ]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("files:get:"))
async def send_price_file(callback: CallbackQuery) -> None:
    try:
        _, _, warehouse, version = callback.data.split(":", 3)
    except ValueError:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    if warehouse not in WAREHOUSES or version not in ("base", "10"):
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    source = price_storage_path / f"{warehouse}.xlsx"
    if not source.exists():
        await callback.answer("Исходный файл не найден. Попросите администратора обновить прайс.", show_alert=True)
        return
    await callback.answer("Готовлю файл…")
    warehouse_name = WAREHOUSES[warehouse]
    if version == "base":
        await callback.message.answer_document(
            FSInputFile(source, filename=f"прайс {warehouse_name} базовый.xlsx"),
            caption=f"📄 <b>{warehouse_name}</b> · базовые цены",
        )
        return
    try:
        with tempfile.TemporaryDirectory() as temp_name:
            destination = Path(temp_name) / f"прайс {warehouse_name} скидка 10.xlsx"
            changed = await asyncio.to_thread(generate_discounted_price, source, destination, 10)
            await callback.message.answer_document(
                FSInputFile(destination, filename=destination.name),
                caption=(
                    f"📄 <b>{warehouse_name}</b> · цены со скидкой 10%\n\n"
                    f"Пересчитано товарных позиций: <b>{changed}</b>"
                ),
            )
    except Exception:
        logging.exception("Не удалось сформировать прайс со скидкой для %s", warehouse)
        await callback.message.answer(
            "⚠️ Не удалось сформировать файл. Попробуйте ещё раз или сообщите администратору."
        )


@router.callback_query(F.data == "main:prices")
async def open_price_search(callback: CallbackQuery, state: FSMContext) -> None:
    if not prices_db.import_statuses():
        await callback.answer("Прайсы пока не загружены.", show_alert=True)
        return
    await state.set_state(AppState.price_search)
    await state.update_data(price_result_ids=[], price_result_page=0)
    await callback.message.edit_text(
        "💰 <b>Цены и наличие</b>\n\n"
        "Введите название товарной группы. Можно использовать бренд, модель, категорию "
        "или их часть — точное совпадение не требуется.\n\n"
        "Например: <code>OGGO VLIQ</code> или <code>Dojo 12000</code>",
        reply_markup=back_main(),
    )
    await callback.answer()


@router.message(AppState.price_search, F.text)
async def search_prices(message: Message, state: FSMContext) -> None:
    query = message.text.strip()
    if len(query) > 200:
        await message.answer("⚠️ Запрос слишком длинный. Укажите только товарную группу.", reply_markup=back_main())
        return
    normalized_query = normalize_price_text(query)
    has_specification = any(char.isdigit() for char in normalized_query)
    item_results = prices_db.search_items(query) if has_specification else []
    # Числа в запросе обычно означают сопротивление, объём или модель. В таком
    # случае показываем конкретные позиции, но не используем товарный код в поиске.
    specific_item_search = has_specification and 1 <= len(item_results) <= 20
    if specific_item_search:
        item_ids = [item.callback_id for item in item_results]
        await state.update_data(
            price_item_result_ids=item_ids,
            price_item_result_page=0,
            price_result_kind="items",
        )
        await message.answer(
            price_item_search_text(len(item_ids), 0),
            reply_markup=price_item_search_keyboard(item_ids, 0),
        )
        return

    groups = prices_db.search_groups(query)
    if not groups:
        await message.answer(
            "🤷 <b>Товарная группа не найдена</b>\n\n"
            "Попробуйте сократить запрос, проверить название бренда или указать модель.",
            reply_markup=back_main(),
        )
        return
    group_ids = [group.callback_id for group in groups]
    await state.update_data(price_result_ids=group_ids, price_result_page=0, price_result_kind="groups")
    await message.answer(
        price_search_text(len(groups), 0),
        reply_markup=price_search_keyboard(group_ids, 0),
    )


@router.callback_query(F.data.startswith("price:r:"))
async def change_price_results_page(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        page = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректная страница.", show_alert=True)
        return
    group_ids = (await state.get_data()).get("price_result_ids", [])
    if not group_ids:
        await callback.answer("Выполните поиск ещё раз.", show_alert=True)
        return
    await state.update_data(price_result_page=page)
    await callback.message.edit_text(
        price_search_text(len(group_ids), page),
        reply_markup=price_search_keyboard(group_ids, page),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("price:ir:"))
async def change_price_item_results_page(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        page = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректная страница.", show_alert=True)
        return
    item_ids = (await state.get_data()).get("price_item_result_ids", [])
    if not item_ids:
        await callback.answer("Выполните поиск ещё раз.", show_alert=True)
        return
    await state.update_data(price_item_result_page=page)
    await callback.message.edit_text(
        price_item_search_text(len(item_ids), page),
        reply_markup=price_item_search_keyboard(item_ids, page),
    )
    await callback.answer()


@router.message(AppState.price_search)
async def price_non_text(message: Message) -> None:
    await message.answer("Отправьте название товара обычным текстом.", reply_markup=back_main())


@router.callback_query(F.data.startswith("price:g:"))
async def show_price_group(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        callback_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    if not await render_price_group(callback.message, callback_id, state):
        await callback.answer("Прайс обновился. Выполните поиск ещё раз.", show_alert=True)
        return
    await callback.answer()


async def render_price_group(message: Message, callback_id: int, state: FSMContext) -> bool:
    details = prices_db.group_details(callback_id)
    if not details:
        return False
    selected_ids = set((await state.get_data()).get("price_selected_ids", []))
    selected = callback_id in selected_ids
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📋 Позиции", callback_data=f"price:v:{callback_id}:0"),
            InlineKeyboardButton(
                text="✅ В подборке" if selected else "➕ В подборку",
                callback_data=f"price:remove:{callback_id}" if selected else f"price:add:{callback_id}",
            ),
        ],
        compact_nav("price:back", search_data="main:prices"),
    ])
    text = format_price_group(details)
    if len(text) <= 4000:
        await message.edit_text(text, reply_markup=keyboard)
    else:
        await message.edit_text(
            f"💰 <b>{html.escape(details.summary.display_name)}</b>\n\n"
            "В группе много ценовых уровней — отправляю подробный расчёт отдельным сообщением.",
            reply_markup=keyboard,
        )
        await message.answer(text)
    return True


@router.callback_query(F.data == "price:back")
async def back_to_price_results(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    group_ids = data.get("price_result_ids", [])
    page = data.get("price_result_page", 0)
    if not group_ids:
        await callback.answer("Результаты поиска уже недоступны. Выполните поиск ещё раз.", show_alert=True)
        return
    await callback.message.edit_text(
        price_search_text(len(group_ids), page),
        reply_markup=price_search_keyboard(group_ids, page),
    )
    await callback.answer()


@router.callback_query(F.data == "price:item_back")
async def back_to_price_item_results(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    item_ids = data.get("price_item_result_ids", [])
    page = data.get("price_item_result_page", 0)
    if not item_ids:
        await callback.answer("Результаты поиска уже недоступны. Выполните поиск ещё раз.", show_alert=True)
        return
    await callback.message.edit_text(
        price_item_search_text(len(item_ids), page),
        reply_markup=price_item_search_keyboard(item_ids, page),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("price:i:"))
async def show_price_item(callback: CallbackQuery) -> None:
    try:
        callback_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный товар.", show_alert=True)
        return
    item = prices_db.item_details(callback_id)
    if not item:
        await callback.answer("Прайс обновился. Выполните поиск ещё раз.", show_alert=True)
        return
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        compact_nav("price:item_back", search_data="main:prices"),
    ])
    await callback.message.edit_text(format_price_item(item), reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("price:add:"))
async def add_price_group(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        callback_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный товар.", show_alert=True)
        return
    data = await state.get_data()
    selected = list(dict.fromkeys([*data.get("price_selected_ids", []), callback_id]))
    if len(selected) > MAX_SELECTED_PRICE_GROUPS:
        await callback.answer(f"В подборку можно добавить до {MAX_SELECTED_PRICE_GROUPS} групп.", show_alert=True)
        return
    await state.update_data(price_selected_ids=selected)
    await callback.answer(f"Добавлено в подборку: {len(selected)}")
    await render_price_group(callback.message, callback_id, state)


@router.callback_query(F.data.startswith("price:remove:"))
async def remove_price_group(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        callback_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный товар.", show_alert=True)
        return
    data = await state.get_data()
    selected = [value for value in data.get("price_selected_ids", []) if value != callback_id]
    await state.update_data(price_selected_ids=selected)
    await callback.answer("Удалено из подборки")
    await render_price_group(callback.message, callback_id, state)


@router.callback_query(F.data == "price:add_all")
async def add_all_price_results(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    result_ids = data.get("price_result_ids", [])
    selected = list(dict.fromkeys([*data.get("price_selected_ids", []), *result_ids]))
    truncated = len(selected) > MAX_SELECTED_PRICE_GROUPS
    selected = selected[:MAX_SELECTED_PRICE_GROUPS]
    await state.update_data(price_selected_ids=selected)
    message = f"Добавлено в подборку: {len(selected)}"
    if truncated:
        message += f" (достигнут лимит {MAX_SELECTED_PRICE_GROUPS})"
    await callback.answer(message, show_alert=True)


def price_cart_keyboard(selected_ids: list[int]) -> InlineKeyboardMarkup:
    summaries = {group.callback_id: group for group in prices_db.group_summaries()}
    rows = [
        [InlineKeyboardButton(
            text=f"❌ {price_group_label(summaries[value].display_name, summaries[value].category_name)}",
            callback_data=f"price:cart_remove:{value}",
        )]
        for value in selected_ids if value in summaries
    ]
    if rows:
        rows.extend([
            [
                InlineKeyboardButton(text="📄 Базовый", callback_data="price:export:0"),
                InlineKeyboardButton(text="📄 Скидка −10%", callback_data="price:export:10"),
            ],
            [InlineKeyboardButton(text="🗑 Очистить подборку", callback_data="price:cart_clear")],
        ])
    rows.append(compact_nav("price:back", search_data="main:prices"))
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "price:cart")
async def show_price_cart(callback: CallbackQuery, state: FSMContext) -> None:
    selected = (await state.get_data()).get("price_selected_ids", [])
    text = (
        f"🧺 <b>Подборка для прайса</b>\n\nВыбрано товарных групп: <b>{len(selected)}</b>\n\n"
        "Нажмите на товар, чтобы удалить его, либо сформируйте Excel-файл."
        if selected else
        "🧺 <b>Подборка пока пуста</b>\n\nДобавьте одну группу из карточки товара или все результаты поиска целиком."
    )
    await callback.message.edit_text(text, reply_markup=price_cart_keyboard(selected))
    await callback.answer()


@router.callback_query(F.data.startswith("price:cart_remove:"))
async def remove_from_price_cart(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        callback_id = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректный товар.", show_alert=True)
        return
    data = await state.get_data()
    await state.update_data(price_selected_ids=[v for v in data.get("price_selected_ids", []) if v != callback_id])
    selected = (await state.get_data()).get("price_selected_ids", [])
    await callback.message.edit_text(
        f"🧺 <b>Подборка для прайса</b>\n\nВыбрано товарных групп: <b>{len(selected)}</b>",
        reply_markup=price_cart_keyboard(selected),
    )
    await callback.answer("Удалено")


@router.callback_query(F.data == "price:cart_clear")
async def clear_price_cart(callback: CallbackQuery, state: FSMContext) -> None:
    await state.update_data(price_selected_ids=[])
    await callback.message.edit_text(
        "🧺 <b>Подборка пока пуста</b>\n\nДобавьте одну группу из карточки товара или все результаты поиска целиком.",
        reply_markup=price_cart_keyboard([]),
    )
    await callback.answer("Подборка очищена")


@router.callback_query(F.data.startswith("price:export:"))
async def export_selected_prices(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        discount = int(callback.data.rsplit(":", 1)[1])
    except ValueError:
        await callback.answer("Некорректная скидка.", show_alert=True)
        return
    selected = (await state.get_data()).get("price_selected_ids", [])
    if not selected:
        await callback.answer("Сначала добавьте товары в подборку.", show_alert=True)
        return
    await callback.answer("Формирую прайс…")
    try:
        with tempfile.TemporaryDirectory() as temp_name:
            suffix = "скидка 10" if discount else "базовые цены"
            merge_keys, availability = prices_db.selection_availability(selected)
            sent = 0
            for warehouse in ("center", "west", "ural"):
                source = price_storage_path / f"{warehouse}.xlsx"
                if not source.exists():
                    continue
                destination = Path(temp_name) / f"прайс {WAREHOUSES[warehouse]} подборка {suffix}.xlsx"
                try:
                    count = await asyncio.to_thread(
                        generate_selected_price, source, destination, merge_keys, availability, discount
                    )
                except ValueError as error:
                    if "отсутствуют" in str(error):
                        continue
                    raise
                await callback.message.answer_document(
                    FSInputFile(destination, filename=destination.name),
                    caption=(f"📄 <b>{WAREHOUSES[warehouse]}</b> · выбранные товары\n\n"
                             f"Товарных позиций: <b>{count}</b> · "
                             f"{'скидка 10%' if discount else 'базовые цены'}"),
                )
                sent += 1
            if not sent:
                raise ValueError("Выбранные товары отсутствуют в действующих прайсах.")
    except Exception:
        logging.exception("Не удалось сформировать прайс по подборке")
        await callback.message.answer("⚠️ Не удалось сформировать файл. Попробуйте ещё раз.")


@router.callback_query(F.data.startswith("price:v:"))
async def show_price_variants(callback: CallbackQuery) -> None:
    try:
        _, _, raw_id, raw_page = callback.data.split(":", 3)
        callback_id, page = int(raw_id), int(raw_page)
    except ValueError:
        await callback.answer("Некорректный запрос.", show_alert=True)
        return
    result = prices_db.group_variants(callback_id)
    if not result:
        await callback.answer("Прайс обновился. Выполните поиск ещё раз.", show_alert=True)
        return
    summary, variants = result
    await callback.message.edit_text(
        format_price_variants(summary, variants, page),
        reply_markup=price_variants_keyboard(callback_id, page, len(variants)),
    )
    await callback.answer()


def products_keyboard(admin: bool = False) -> InlineKeyboardMarkup:
    products = materials_db.list_products(visible_only=not admin)
    rows = []
    for product in products:
        marker = "🟢" if product.is_visible else "⚪️"
        text = f"{marker} {product.name}" if admin else product.name
        prefix = "adm:p" if admin else "db:p"
        rows.append([InlineKeyboardButton(text=text, callback_data=f"{prefix}:{product.id}")])
    if admin:
        rows.extend(button_grid([
            InlineKeyboardButton(text="➕ Товар", callback_data="adm:add_product"),
            InlineKeyboardButton(text="💰 Прайсы", callback_data="adm:prices"),
            InlineKeyboardButton(text="👥 Доступ", callback_data="adm:access"),
            InlineKeyboardButton(text="📊 Excel", callback_data="adm:excel"),
            InlineKeyboardButton(text="✏️ Текст /прайс", callback_data="adm:price_message"),
        ]))
        rows.append([InlineKeyboardButton(text="💾 Скачать резервную копию", callback_data="adm:backup")])
    rows.append(compact_nav())
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "main:database")
async def open_database(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    products = materials_db.list_products(visible_only=True)
    text = (
        "🗃 <b>База данных</b>\n\nВыберите продукцию, чтобы посмотреть доступные материалы:"
        if products
        else "🗃 <b>База данных</b>\n\nМатериалы пока не опубликованы. Загляните сюда позднее."
    )
    await callback.message.edit_text(text, reply_markup=products_keyboard())
    await callback.answer()


@router.callback_query(F.data.startswith("db:p:"))
async def open_product(callback: CallbackQuery) -> None:
    product = materials_db.get_product(int(callback.data.rsplit(":", 1)[1]))
    if not product or not product.is_visible:
        await callback.answer("Товар недоступен.", show_alert=True)
        return
    sections = materials_db.list_sections(product.id)
    rows = [[InlineKeyboardButton(text=s.name, callback_data=f"db:s:{s.id}")] for s in sections]
    rows.append(compact_nav("main:database"))
    text = (
        f"📦 <b>{html.escape(product.name)}</b>\n\nВыберите нужный раздел:"
        if sections
        else f"📦 <b>{html.escape(product.name)}</b>\n\nВ этом разделе пока нет опубликованных материалов."
    )
    await callback.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await callback.answer()


async def send_material(bot: Bot, chat_id: int, material: Material) -> None:
    if material.kind == "text":
        await bot.send_message(chat_id, material.text or "", parse_mode=None)
    elif material.kind == "photo":
        await bot.send_photo(chat_id, material.file_id, caption=material.caption, parse_mode=None)
    elif material.kind == "document":
        await bot.send_document(chat_id, material.file_id, caption=material.caption, parse_mode=None)


@router.callback_query(F.data.startswith("db:s:"))
async def deliver_section(callback: CallbackQuery, bot: Bot) -> None:
    section = materials_db.get_section(int(callback.data.rsplit(":", 1)[1]))
    if not section:
        await callback.answer("Раздел не найден.", show_alert=True)
        return
    product = materials_db.get_product(section.product_id)
    if not product or not product.is_visible:
        await callback.answer("Материал недоступен.", show_alert=True)
        return
    items = materials_db.list_materials(section.id)
    if not items:
        await callback.answer("В этом разделе пока нет опубликованных материалов.", show_alert=True)
        return
    await callback.answer("Материалы отправляются…")
    await bot.send_message(
        callback.message.chat.id,
        f"📎 <b>{html.escape(product.name)}</b>\nРаздел: <b>{html.escape(section.name)}</b>",
    )
    for item in items:
        try:
            await send_material(bot, callback.message.chat.id, item)
        except Exception:
            logging.exception("Не удалось отправить материал %s", item.id)
            await bot.send_message(
                callback.message.chat.id,
                "⚠️ Один из материалов временно недоступен. Сообщите об этом администратору.",
            )


async def require_admin(callback: CallbackQuery) -> bool:
    if is_admin(callback.from_user.id) and callback.message.chat.type == "private":
        return True
    await callback.answer("Этот раздел доступен только администратору.", show_alert=True)
    return False


async def require_broadcaster(callback: CallbackQuery) -> bool:
    if can_broadcast(callback.from_user.id, callback.from_user.username) and callback.message.chat.type == "private":
        return True
    await callback.answer("Раздел доступен администраторам рассылок.", show_alert=True)
    return False


async def require_price_message_admin(callback: CallbackQuery) -> bool:
    if can_manage_price_message(
        callback.from_user.id, callback.from_user.username
    ) and callback.message.chat.type == "private":
        return True
    await callback.answer("Эта настройка доступна администраторам.", show_alert=True)
    return False


async def cleanup_pending_excel(state: FSMContext) -> None:
    data = await state.get_data()
    for key in ("pending_excel", "pending_price"):
        pending = data.get(key)
        if pending:
            Path(pending).unlink(missing_ok=True)


def build_backup_archive(archive_path: Path) -> None:
    with tempfile.TemporaryDirectory() as temp_name:
        temp_dir = Path(temp_name)
        database_copy = temp_dir / "materials.sqlite3"
        excel_copy = temp_dir / "managers.xlsx"
        materials_db.backup_to(database_copy)
        shutil.copy2(active_excel_path, excel_copy)
        prices_database = globals().get("prices_db")
        prices_copy = temp_dir / "prices.sqlite3"
        if prices_database is not None:
            prices_database.backup_to(prices_copy)
        metadata = temp_dir / "README.txt"
        metadata.write_text(
            "Резервная копия Ural Vape Regions Bot\n"
            f"Создана: {datetime.now().astimezone().isoformat(timespec='seconds')}\n"
            f"Записей территорий: {len(catalog.entries)}\n"
            "materials.sqlite3 — товары, разделы и материалы\n"
            "managers.xlsx — действующая таблица территорий\n"
            "prices.sqlite3 — загруженные складские цены и товарные группы\n"
            "price_files/ — последние исходные прайсы складов\n",
            encoding="utf-8",
        )
        with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(database_copy, database_copy.name)
            archive.write(excel_copy, excel_copy.name)
            if prices_copy.exists():
                archive.write(prices_copy, prices_copy.name)
            storage = globals().get("price_storage_path")
            if storage and storage.exists():
                for warehouse in WAREHOUSES:
                    source = storage / f"{warehouse}.xlsx"
                    if source.exists():
                        archive.write(source, f"price_files/{source.name}")
            archive.write(metadata, metadata.name)


@router.callback_query(F.data == "adm:backup")
async def download_backup(callback: CallbackQuery) -> None:
    if not await require_admin(callback):
        return
    await callback.answer("Готовлю резервную копию…")
    await callback.message.answer(
        "⏳ <b>Создаю резервную копию</b>\n\n"
        "Это может занять несколько секунд. Не закрывайте чат."
    )
    try:
        with tempfile.TemporaryDirectory() as temp_name:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            archive_path = Path(temp_name) / f"ural-vape-bot-backup_{timestamp}.zip"
            await asyncio.to_thread(build_backup_archive, archive_path)
            await callback.message.answer_document(
                FSInputFile(archive_path),
                caption=(
                    "✅ <b>Резервная копия готова</b>\n\n"
                    "В архиве находятся база материалов, таблица территорий и складские прайсы. "
                    "Храните файл в надёжном месте."
                ),
            )
    except Exception:
        logging.exception("Не удалось создать резервную копию")
        await callback.message.answer(
            "⚠️ Не удалось создать резервную копию. Попробуйте ещё раз или проверьте журнал сервера."
        )


@router.callback_query(F.data == "adm:excel")
async def start_excel_upload(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await cleanup_pending_excel(state)
    await state.set_state(AdminState.excel_upload)
    await callback.message.edit_text(
        "📊 <b>Обновление таблицы территорий</b>\n\n"
        "Отправьте Excel-файл в формате <code>.xlsx</code>. В первой строке должны быть колонки:\n\n"
        "• <b>Местоположение</b>\n"
        "• <b>Территория</b>\n"
        "• <b>Менеджер</b>\n\n"
        "Сначала бот проверит файл и покажет сводку. Действующая таблица не изменится без подтверждения.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="❌ Отмена", callback_data="main:admin")]]
        ),
    )
    await callback.answer()


@router.message(AdminState.excel_upload)
async def receive_excel(message: Message, state: FSMContext, bot: Bot) -> None:
    if not is_admin(message.from_user.id) or message.chat.type != "private":
        return
    if not message.document or not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer(
            "⚠️ Нужен документ в формате <code>.xlsx</code>. Отправьте правильный файл или нажмите /cancel."
        )
        return
    pending_dir = managed_excel_path.parent / "pending"
    pending_dir.mkdir(parents=True, exist_ok=True)
    pending_path = pending_dir / f"excel_{message.document.file_unique_id}.xlsx"
    await message.answer("⏳ Файл получен. Проверяю структуру и данные…")
    try:
        await bot.download(message.document.file_id, destination=pending_path)
        checked_catalog, details = await asyncio.to_thread(validate_excel, pending_path)
    except Exception as error:
        pending_path.unlink(missing_ok=True)
        logging.warning("Отклонён Excel от администратора: %s", error)
        await message.answer(
            "❌ <b>Файл не прошёл проверку</b>\n\n"
            f"Причина: {html.escape(str(error))}\n\n"
            "Действующая таблица не изменена."
        )
        return
    await state.set_state(AdminState.excel_confirmation)
    await state.update_data(
        pending_excel=str(pending_path),
        excel_rows=len(checked_catalog.entries),
        excel_sheet=details[0],
        excel_name=message.document.file_name,
    )
    await message.answer(
        "✅ <b>Файл успешно проверен</b>\n\n"
        f"Файл: <code>{html.escape(message.document.file_name)}</code>\n"
        f"Лист: <b>{html.escape(details[0])}</b>\n"
        f"Корректных записей: <b>{len(checked_catalog.entries):,}</b>\n\n"
        "Применить эту таблицу? Текущая версия будет сохранена в резервную копию.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(text="✅ Применить", callback_data="adm:apply_excel"),
                    InlineKeyboardButton(text="❌ Отмена", callback_data="adm:cancel_excel"),
                ],
            ]
        ),
    )


def apply_pending_excel(pending_path: Path) -> tuple[Catalog, Path]:
    checked_catalog, _ = validate_excel(pending_path)
    backup_dir = managed_excel_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_path = backup_dir / f"managers_{timestamp}.xlsx"
    shutil.copy2(active_excel_path, backup_path)
    managed_excel_path.parent.mkdir(parents=True, exist_ok=True)
    os.replace(pending_path, managed_excel_path)
    checked_catalog.path = managed_excel_path
    return checked_catalog, backup_path


@router.callback_query(F.data == "adm:apply_excel")
async def apply_excel(callback: CallbackQuery, state: FSMContext) -> None:
    global catalog, active_excel_path
    if not await require_admin(callback):
        return
    data = await state.get_data()
    pending_path = Path(data.get("pending_excel", ""))
    if not pending_path.is_file():
        await state.clear()
        await callback.answer("Файл проверки не найден. Загрузите его ещё раз.", show_alert=True)
        return
    await callback.answer("Применяю таблицу…")
    try:
        new_catalog, backup_path = await asyncio.to_thread(apply_pending_excel, pending_path)
        catalog = new_catalog
        active_excel_path = managed_excel_path
    except Exception:
        logging.exception("Не удалось применить новый Excel")
        await callback.message.edit_text(
            "❌ Не удалось применить таблицу. Действующая версия сохранена без изменений.",
            reply_markup=products_keyboard(admin=True),
        )
        return
    await state.clear()
    await callback.message.edit_text(
        "✅ <b>Таблица обновлена</b>\n\n"
        f"Загружено записей: <b>{len(catalog.entries):,}</b>\n"
        "Новые данные уже используются в поиске — перезапуск бота не требуется.\n\n"
        f"Предыдущая версия сохранена: <code>{html.escape(backup_path.name)}</code>",
        reply_markup=products_keyboard(admin=True),
    )


@router.callback_query(F.data == "adm:cancel_excel")
async def cancel_excel(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await cleanup_pending_excel(state)
    await state.clear()
    await callback.message.edit_text(
        "Обновление таблицы отменено. Действующие данные не изменились.",
        reply_markup=products_keyboard(admin=True),
    )
    await callback.answer()


def price_admin_keyboard() -> InlineKeyboardMarkup:
    statuses = prices_db.import_statuses()
    buttons = []
    for warehouse in ("center", "west", "ural"):
        marker = "✅" if warehouse in statuses else "➕"
        buttons.append(InlineKeyboardButton(
                text=f"{marker} {WAREHOUSES[warehouse]}",
                callback_data=f"adm:price_wh:{warehouse}",
            ))
    rows = button_grid(buttons)
    rows.append(compact_nav("main:admin"))
    return InlineKeyboardMarkup(inline_keyboard=rows)


def price_status_text() -> str:
    statuses = prices_db.import_statuses()
    lines = [
        "💰 <b>Управление прайсами</b>",
        "",
        "Выберите склад, для которого хотите загрузить свежий прайс.",
        "Данные остальных складов не изменятся.",
        "",
    ]
    for warehouse in ("center", "west", "ural"):
        status = statuses.get(warehouse)
        if status:
            date = (status["price_date"] or status["updated_at"] or "").split("T", 1)[0]
            lines.append(
                f"✅ <b>{WAREHOUSES[warehouse]}</b> — {status['item_count']} позиций, "
                f"прайс от {html.escape(date)}"
            )
        else:
            lines.append(f"➕ <b>{WAREHOUSES[warehouse]}</b> — прайс не загружен")
    return "\n".join(lines)


@router.callback_query(F.data == "adm:prices")
async def admin_prices(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await cleanup_pending_excel(state)
    await state.clear()
    await callback.message.edit_text(price_status_text(), reply_markup=price_admin_keyboard())
    await callback.answer()


@router.callback_query(F.data.startswith("adm:price_wh:"))
async def start_price_upload(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    warehouse = callback.data.rsplit(":", 1)[1]
    if warehouse not in WAREHOUSES:
        await callback.answer("Неизвестный склад.", show_alert=True)
        return
    await cleanup_pending_excel(state)
    await state.set_state(AdminState.price_upload)
    await state.update_data(price_warehouse=warehouse)
    await callback.message.edit_text(
        f"💰 <b>Обновление прайса</b>\n"
        f"Склад: <b>{WAREHOUSES[warehouse]}</b>\n\n"
        "Отправьте свежий прайс в формате <code>.xlsx</code>. Бот проверит структуру, "
        "товарные группы и цены, после чего покажет сводку перед применением.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="❌ Отмена", callback_data="adm:prices")]]
        ),
    )
    await callback.answer()


@router.message(AdminState.price_upload)
async def receive_price_file(message: Message, state: FSMContext, bot: Bot) -> None:
    if not is_admin(message.from_user.id) or message.chat.type != "private":
        return
    data = await state.get_data()
    warehouse = data.get("price_warehouse")
    if warehouse not in WAREHOUSES:
        await state.clear()
        await message.answer("Склад не выбран. Начните загрузку заново.")
        return
    if not message.document or not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer("⚠️ Отправьте прайс как документ в формате <code>.xlsx</code>.")
        return
    pending_dir = price_storage_path / "pending"
    pending_dir.mkdir(parents=True, exist_ok=True)
    pending_path = pending_dir / f"{warehouse}_{message.document.file_unique_id}.xlsx"
    await message.answer("⏳ Прайс получен. Проверяю группы, цены и ассортимент…")
    try:
        await bot.download(message.document.file_id, destination=pending_path)
        parsed = await asyncio.to_thread(parse_price_file, pending_path)
    except Exception as error:
        pending_path.unlink(missing_ok=True)
        logging.warning("Отклонён прайс склада %s: %s", warehouse, error)
        await message.answer(
            "❌ <b>Прайс не прошёл проверку</b>\n\n"
            f"Причина: {html.escape(str(error))}\n\n"
            "Действующие цены не изменены."
        )
        return
    await state.set_state(AdminState.price_confirmation)
    await state.update_data(
        pending_price=str(pending_path),
        price_warehouse=warehouse,
        price_file_name=message.document.file_name,
    )
    price_date = parsed.price_date.split("T", 1)[0] if parsed.price_date else "не указана"
    await message.answer(
        "✅ <b>Прайс успешно проверен</b>\n\n"
        f"Склад: <b>{WAREHOUSES[warehouse]}</b>\n"
        f"Дата прайса: <b>{html.escape(price_date)}</b>\n"
        f"Товарных групп: <b>{len(parsed.groups)}</b>\n"
        f"Товарных позиций: <b>{parsed.item_count}</b>\n"
        f"Позиций с пометкой «АКЦИЯ»: <b>{parsed.action_count}</b>\n\n"
        "Применить этот прайс? Предыдущая версия выбранного склада будет сохранена.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Применить", callback_data="adm:apply_price"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="adm:cancel_price"),
            ],
        ]),
    )


def apply_pending_price(
    pending_path: Path, warehouse: str, file_name: str
) -> tuple[int, int, Path | None, dict | None]:
    parsed = parse_price_file(pending_path)
    report = prices_db.build_change_report(warehouse, parsed, file_name)
    backup_dir = price_storage_path / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    database_backup = backup_dir / f"prices_before_{warehouse}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.sqlite3"
    prices_db.backup_to(database_backup)
    source_backup = save_price_source(pending_path, price_storage_path, warehouse)
    prices_db.replace_warehouse(warehouse, parsed, file_name)
    prices_db.save_latest_report(report)
    pending_path.unlink(missing_ok=True)
    return len(parsed.groups), parsed.item_count, source_backup, report


@router.callback_query(F.data == "adm:apply_price")
async def apply_price(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    if not await require_admin(callback):
        return
    data = await state.get_data()
    warehouse = data.get("price_warehouse")
    pending_path = Path(data.get("pending_price", ""))
    if warehouse not in WAREHOUSES or not pending_path.is_file():
        await state.clear()
        await callback.answer("Файл проверки не найден. Загрузите прайс ещё раз.", show_alert=True)
        return
    if warehouse in price_updates_in_progress:
        await callback.answer("Прайс этого склада уже обрабатывается. Дождитесь завершения.", show_alert=True)
        return
    price_updates_in_progress.add(warehouse)
    await callback.answer()
    try:
        await edit_or_answer(
            callback.message,
            "⏳ <b>Применяю прайс</b>\n\n"
            f"Склад: <b>{WAREHOUSES[warehouse]}</b>\n\n"
            "Проверяю изменения цен и наличия, сохраняю предыдущую версию и формирую отчёт. "
            "Это может занять некоторое время — повторно нажимать ничего не нужно."
        )
        group_count, item_count, _, report = await asyncio.to_thread(
            apply_pending_price, pending_path, warehouse, data.get("price_file_name", pending_path.name)
        )
    except Exception:
        logging.exception("Не удалось применить прайс склада %s", warehouse)
        await edit_or_answer(
            callback.message,
            "❌ Не удалось применить прайс. Предыдущие данные сохранены.",
            reply_markup=price_admin_keyboard(),
        )
        return
    finally:
        price_updates_in_progress.discard(warehouse)
    await state.clear()
    try:
        wait_matches = await notify_waitlist_matches(bot, {warehouse: report}) if report else 0
    except Exception:
        logging.exception("Не удалось проверить лист ожидания после обновления прайса")
        wait_matches = 0
    if report:
        wait_note = (
            f"\n\n🔔 Уведомлений по листу ожидания: <b>{wait_matches}</b>"
            if wait_matches else ""
        )
        await edit_or_answer(
            callback.message,
            "✅ <b>Прайс обновлён</b>\n\n" + format_price_report(report) + wait_note,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📥 Скачать подробный Excel", callback_data=f"reports:file:{warehouse}")],
                [
                    InlineKeyboardButton(text="📊 Изменения", callback_data="main:price_reports"),
                    InlineKeyboardButton(text="💰 Прайсы", callback_data="adm:prices"),
                ],
                compact_nav(),
            ]),
        )
    else:
        await edit_or_answer(
            callback.message,
            "✅ <b>Прайс обновлён</b>\n\n"
            f"Склад: <b>{WAREHOUSES[warehouse]}</b>\n"
            f"Товарных групп: <b>{group_count}</b>\n"
            f"Товарных позиций: <b>{item_count}</b>\n\n"
            "Это первая загрузка склада, поэтому сравнение появится при следующем обновлении.",
            reply_markup=price_admin_keyboard(),
        )


@router.callback_query(F.data == "adm:cancel_price")
async def cancel_price(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await cleanup_pending_excel(state)
    await state.clear()
    await callback.message.edit_text(
        "Загрузка прайса отменена. Действующие цены не изменились.",
        reply_markup=price_admin_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "main:admin")
async def admin_menu(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_price_message_admin(callback):
        return
    await state.clear()
    if not is_admin(callback.from_user.id):
        await callback.message.edit_text(
            "⚙️ <b>Управление</b>\n\n"
            "Здесь можно настроить сообщение, которое бот отправляет перед актуальными прайсами.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ Текст /прайс", callback_data="adm:price_message")],
                compact_nav(),
            ]),
        )
        await callback.answer()
        return
    await callback.message.edit_text(
        "⚙️ <b>Управление базой</b>\n\n"
        "Создавайте товары, добавляйте разделы и наполняйте их материалами. "
        "Скрытые товары видны здесь, но недоступны пользователям.",
        reply_markup=products_keyboard(admin=True),
    )
    await callback.answer()


def price_message_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Изменить сообщение", callback_data="adm:edit_price_message")],
        [InlineKeyboardButton(text="↩️ Вернуть стандартное", callback_data="adm:reset_price_message")],
        compact_nav("main:admin"),
    ])


async def show_price_message_settings(callback: CallbackQuery) -> None:
    current = materials_db.get_setting(PRICE_MESSAGE_SETTING) or DEFAULT_PRICE_COMMAND_MESSAGE
    await callback.message.edit_text(
        "✏️ <b>Сообщение команды /прайс</b>\n\n"
        "Сейчас перед файлами отправляется:\n\n"
        f"{current}",
        reply_markup=price_message_keyboard(),
    )


@router.callback_query(F.data == "adm:price_message")
async def manage_price_message(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_price_message_admin(callback):
        return
    await state.clear()
    await show_price_message_settings(callback)
    await callback.answer()


@router.callback_query(F.data == "adm:edit_price_message")
async def ask_price_message(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_price_message_admin(callback):
        return
    await state.set_state(AdminState.price_message)
    await callback.message.edit_text(
        "✏️ <b>Новое сообщение для /прайс</b>\n\n"
        "Отправьте текст одним сообщением. Можно использовать обычное форматирование Telegram: "
        "жирный шрифт, курсив и ссылки.\n\n"
        "После сохранения этот текст будут видеть пользователи перед файлами прайсов.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[compact_nav("adm:price_message")]),
    )
    await callback.answer()


@router.message(AdminState.price_message, F.text)
async def save_price_message(message: Message, state: FSMContext) -> None:
    if not message.from_user or not can_manage_price_message(
        message.from_user.id, message.from_user.username
    ):
        await state.clear()
        return
    if len(message.text.strip()) > 3500:
        await message.answer("Сообщение слишком длинное. Сократите его до 3500 символов.")
        return
    materials_db.set_setting(PRICE_MESSAGE_SETTING, message.html_text.strip())
    await state.clear()
    await message.answer(
        "✅ <b>Сообщение для /прайс обновлено</b>\n\n"
        "Новый текст будет использоваться при следующем запросе.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="👁 Посмотреть", callback_data="adm:price_message")],
            compact_nav(),
        ]),
    )


@router.callback_query(F.data == "adm:reset_price_message")
async def reset_price_message(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_price_message_admin(callback):
        return
    materials_db.set_setting(PRICE_MESSAGE_SETTING, DEFAULT_PRICE_COMMAND_MESSAGE)
    await state.clear()
    await show_price_message_settings(callback)
    await callback.answer("Стандартное сообщение восстановлено")


def access_user_label(user) -> str:
    parts = []
    if user.username:
        parts.append(f"@{user.username}")
    if user.telegram_id:
        parts.append(f"ID {user.telegram_id}")
    role = "младший администратор" if user.role == "junior_admin" else "пользователь"
    return f"{' · '.join(parts)} · {role}"


def access_keyboard() -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(
                text=f"👤 {access_user_label(user)}",
                callback_data=f"adm:access_user:{user.id}",
            )
        ]
        for user in materials_db.list_access_users()
    ]
    rows.extend([
        [InlineKeyboardButton(text="➕ Добавить пользователя", callback_data="adm:add_access")],
        compact_nav("main:admin"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "adm:access")
async def manage_access(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await state.clear()
    users = materials_db.list_access_users()
    await callback.message.edit_text(
        "👥 <b>Белый список</b>\n\n"
        f"Пользователей с доступом: <b>{len(users)}</b>\n\n"
        "Добавьте Telegram ID или @username. Username будет привязан к постоянному ID "
        "при первом обращении пользователя к боту.\n\n"
        "Нажмите пользователя, чтобы изменить его роль или закрыть доступ.",
        reply_markup=access_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("adm:access_user:"))
async def manage_access_user(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    access_id = int(callback.data.rsplit(":", 1)[1])
    user = materials_db.get_access_user(access_id)
    if not user:
        await callback.answer("Пользователь не найден.", show_alert=True); return
    next_role = "user" if user.role == "junior_admin" else "junior_admin"
    role_text = "Сделать пользователем" if next_role == "user" else "Назначить младшим администратором"
    await callback.message.edit_text(
        f"👤 <b>{html.escape(access_user_label(user))}</b>\n\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🔑 {role_text}", callback_data=f"adm:set_role:{access_id}:{next_role}")],
            [InlineKeyboardButton(text="🗑 Закрыть доступ", callback_data=f"adm:confirm_access:{access_id}")],
            compact_nav("adm:access"),
        ]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("adm:set_role:"))
async def set_access_role(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    _, _, raw_id, role = callback.data.split(":", 3)
    materials_db.set_access_role(int(raw_id), role)
    await callback.message.edit_text("✅ Роль пользователя обновлена.", reply_markup=access_keyboard())
    await callback.answer()


@router.callback_query(F.data == "adm:add_access")
async def ask_access_user(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await state.set_state(AdminState.access_user)
    await callback.message.edit_text(
        "➕ <b>Новый пользователь</b>\n\n"
        "Отправьте один идентификатор:\n\n"
        "• числовой Telegram ID, например <code>5533726476</code>;\n"
        "• username с символом @, например <code>@username</code>.\n\n"
        "ID надёжнее, потому что username пользователь может изменить.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="❌ Отмена", callback_data="adm:access")]]
        ),
    )
    await callback.answer()


@router.message(AdminState.access_user, F.text)
async def save_access_user(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    try:
        user = materials_db.add_access_user(message.text)
    except ValueError as error:
        await message.answer(f"⚠️ {html.escape(str(error))}")
        return
    except sqlite3.IntegrityError:
        await message.answer("⚠️ Такой пользователь уже находится в белом списке.")
        return
    await state.clear()
    await message.answer(
        f"✅ Доступ предоставлен: <b>{html.escape(access_user_label(user))}</b>",
        reply_markup=access_keyboard(),
    )


@router.callback_query(F.data.startswith("adm:confirm_access:"))
async def confirm_access_delete(callback: CallbackQuery) -> None:
    if not await require_admin(callback):
        return
    access_id = int(callback.data.rsplit(":", 1)[1])
    user = materials_db.get_access_user(access_id)
    if not user:
        await callback.answer("Пользователь уже удалён.", show_alert=True)
        return
    await callback.message.edit_text(
        "⚠️ <b>Закрыть доступ?</b>\n\n"
        f"Пользователь: <b>{html.escape(access_user_label(user))}</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Да, закрыть доступ", callback_data=f"adm:delete_access:{access_id}")],
            [InlineKeyboardButton(text="Отмена", callback_data="adm:access")],
        ]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("adm:delete_access:"))
async def delete_access_user(callback: CallbackQuery) -> None:
    if not await require_admin(callback):
        return
    materials_db.delete_access_user(int(callback.data.rsplit(":", 1)[1]))
    await callback.message.edit_text(
        "✅ Доступ пользователя закрыт.",
        reply_markup=access_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data == "adm:add_product")
async def admin_add_product(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback):
        return
    await state.set_state(AdminState.product_name)
    await callback.message.edit_text(
        "➕ <b>Новый товар</b>\n\nВведите название так, как его должны видеть пользователи:",
        reply_markup=back_main(),
    )
    await callback.answer()


@router.message(AdminState.product_name, F.text)
async def save_product(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    try:
        product = materials_db.add_product(message.text)
    except sqlite3.IntegrityError:
        await message.answer("⚠️ Товар с таким названием уже существует. Введите другое название.")
        return
    await state.clear()
    await message.answer(
        f"✅ Товар <b>{html.escape(product.name)}</b> создан.\n\nТеперь добавьте в него первый раздел.",
        reply_markup=admin_product_keyboard(product.id),
    )


def admin_product_keyboard(product_id: int) -> InlineKeyboardMarkup:
    product = materials_db.get_product(product_id)
    sections = materials_db.list_sections(product_id)
    rows = [[InlineKeyboardButton(text=f"📁 {s.name}", callback_data=f"adm:s:{s.id}")] for s in sections]
    rows.extend([
        [
            InlineKeyboardButton(text="➕ Раздел", callback_data=f"adm:add_section:{product_id}"),
            InlineKeyboardButton(text="✏️ Название", callback_data=f"adm:rename_product:{product_id}"),
        ],
        [
            InlineKeyboardButton(text="🙈 Скрыть" if product and product.is_visible else "👁 Показать", callback_data=f"adm:toggle_product:{product_id}"),
            InlineKeyboardButton(text="🗑 Удалить", callback_data=f"adm:confirm_product:{product_id}"),
        ],
        compact_nav("main:admin"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data.startswith("adm:p:"))
async def admin_product(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    product_id = int(callback.data.rsplit(":", 1)[1])
    product = materials_db.get_product(product_id)
    if not product:
        await callback.answer("Товар не найден.", show_alert=True); return
    status = "доступен пользователям" if product.is_visible else "скрыт"
    await callback.message.edit_text(
        f"📦 <b>{html.escape(product.name)}</b>\n\nСтатус: <b>{status}</b>\n"
        f"Разделов: <b>{len(materials_db.list_sections(product_id))}</b>",
        reply_markup=admin_product_keyboard(product_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("adm:add_section:"))
async def admin_add_section(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback): return
    product_id = int(callback.data.rsplit(":", 1)[1])
    await state.set_state(AdminState.section_name)
    await state.update_data(product_id=product_id)
    await callback.message.edit_text(
        "➕ <b>Новый раздел</b>\n\n"
        "Введите название, например: <code>Декларации</code>, <code>Мокапы</code> или <code>КП</code>.",
        reply_markup=back_main(),
    )
    await callback.answer()


@router.message(AdminState.section_name, F.text)
async def save_section(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id): return
    product_id = (await state.get_data())["product_id"]
    try:
        section = materials_db.add_section(product_id, message.text)
    except sqlite3.IntegrityError:
        await message.answer("⚠️ Раздел с таким названием уже существует."); return
    await state.clear()
    await message.answer(
        f"✅ Раздел <b>{html.escape(section.name)}</b> создан.\n\nТеперь можно добавить материалы.",
        reply_markup=admin_section_keyboard(section.id),
    )


def material_title(material: Material, number: int) -> str:
    if material.kind == "text":
        preview = (material.text or "").replace("\n", " ")[:30]
        return f"🗑 {number}. Текст: {preview}"
    if material.kind == "photo":
        return f"🗑 {number}. Изображение"
    return f"🗑 {number}. {material.file_name or 'Файл'}"


def admin_section_keyboard(section_id: int) -> InlineKeyboardMarkup:
    section = materials_db.get_section(section_id)
    items = materials_db.list_materials(section_id)
    rows = [[InlineKeyboardButton(text=material_title(m, i), callback_data=f"adm:confirm_material:{m.id}")] for i, m in enumerate(items, 1)]
    rows.extend([
        [
            InlineKeyboardButton(text="➕ Материалы", callback_data=f"adm:add_material:{section_id}"),
            InlineKeyboardButton(text="👁 Просмотреть", callback_data=f"adm:preview:{section_id}"),
        ],
        [
            InlineKeyboardButton(text="✏️ Название", callback_data=f"adm:rename_section:{section_id}"),
            InlineKeyboardButton(text="🗑 Удалить", callback_data=f"adm:confirm_section:{section_id}"),
        ],
        compact_nav(f"adm:p:{section.product_id}"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data.startswith("adm:s:"))
async def admin_section(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1])
    section = materials_db.get_section(section_id)
    if not section:
        await callback.answer("Раздел не найден.", show_alert=True); return
    count = len(materials_db.list_materials(section_id))
    await callback.message.edit_text(
        f"📁 <b>{html.escape(section.name)}</b>\n\nМатериалов внутри: <b>{count}</b>\n"
        "Нажмите на материал с символом 🗑, чтобы удалить его.",
        reply_markup=admin_section_keyboard(section_id),
    )
    await callback.answer()


def upload_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Завершить", callback_data="adm:finish_upload"),
            InlineKeyboardButton(text="❌ Отменить", callback_data="main:admin"),
        ],
    ])


@router.callback_query(F.data.startswith("adm:add_material:"))
async def start_upload(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1])
    await state.set_state(AdminState.material_upload)
    await state.update_data(section_id=section_id)
    await callback.message.edit_text(
        "📎 <b>Добавление материалов</b>\n\n"
        "Отправляйте текст, изображения или документы по одному сообщению. "
        "Пользователь получит их в том же порядке.\n\n"
        "Когда всё будет добавлено, нажмите <b>«Завершить»</b>.",
        reply_markup=upload_keyboard(),
    )
    await callback.answer()


@router.message(AdminState.material_upload)
async def save_material(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id): return
    section_id = (await state.get_data())["section_id"]
    if message.text:
        materials_db.add_material(section_id, "text", text=message.text)
        label = "Текст"
    elif message.photo:
        materials_db.add_material(section_id, "photo", file_id=message.photo[-1].file_id, caption=message.caption)
        label = "Изображение"
    elif message.document:
        materials_db.add_material(
            section_id, "document", file_id=message.document.file_id,
            caption=message.caption, file_name=message.document.file_name,
        )
        label = "Файл"
    else:
        await message.answer(
            "⚠️ Этот формат пока не поддерживается. Отправьте текст, изображение или документ.",
            reply_markup=upload_keyboard(),
        ); return
    await message.answer(f"✅ {label} добавлен. Можно отправить следующий материал.", reply_markup=upload_keyboard())


@router.callback_query(F.data == "adm:finish_upload")
async def finish_upload(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback): return
    section_id = (await state.get_data()).get("section_id")
    await state.clear()
    if not section_id or not materials_db.get_section(section_id):
        await callback.message.edit_text("Раздел не найден.", reply_markup=products_keyboard(admin=True))
    else:
        await callback.message.edit_text(
            "✅ <b>Материалы сохранены</b>\n\nРаздел готов к использованию.",
            reply_markup=admin_section_keyboard(section_id),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("adm:preview:"))
async def preview_section(callback: CallbackQuery, bot: Bot) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1])
    items = materials_db.list_materials(section_id)
    if not items:
        await callback.answer("Материалов пока нет.", show_alert=True); return
    await callback.answer("Отправляю предпросмотр…")
    for item in items:
        await send_material(bot, callback.message.chat.id, item)


@router.callback_query(F.data.startswith("adm:rename_product:"))
async def ask_product_rename(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback): return
    product_id = int(callback.data.rsplit(":", 1)[1])
    await state.set_state(AdminState.product_rename); await state.update_data(product_id=product_id)
    await callback.message.edit_text("✏️ <b>Переименование товара</b>\n\nВведите новое название:", reply_markup=back_main()); await callback.answer()


@router.message(AdminState.product_rename, F.text)
async def save_product_rename(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id): return
    product_id = (await state.get_data())["product_id"]
    try: materials_db.rename_product(product_id, message.text)
    except sqlite3.IntegrityError:
        await message.answer("Такое название уже используется."); return
    await state.clear(); await message.answer("✅ Название товара обновлено.", reply_markup=admin_product_keyboard(product_id))


@router.callback_query(F.data.startswith("adm:rename_section:"))
async def ask_section_rename(callback: CallbackQuery, state: FSMContext) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1])
    await state.set_state(AdminState.section_rename); await state.update_data(section_id=section_id)
    await callback.message.edit_text("✏️ <b>Переименование раздела</b>\n\nВведите новое название:", reply_markup=back_main()); await callback.answer()


@router.message(AdminState.section_rename, F.text)
async def save_section_rename(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id): return
    section_id = (await state.get_data())["section_id"]
    try: materials_db.rename_section(section_id, message.text)
    except sqlite3.IntegrityError:
        await message.answer("Такое название уже используется."); return
    await state.clear(); await message.answer("✅ Название раздела обновлено.", reply_markup=admin_section_keyboard(section_id))


@router.callback_query(F.data.startswith("adm:toggle_product:"))
async def toggle_product(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    product_id = int(callback.data.rsplit(":", 1)[1]); materials_db.toggle_product(product_id)
    await callback.message.edit_reply_markup(reply_markup=admin_product_keyboard(product_id)); await callback.answer("Статус изменён")


def confirm_keyboard(yes_data: str, back_data: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Да", callback_data=yes_data),
            InlineKeyboardButton(text="❌ Нет", callback_data=back_data),
        ],
    ])


@router.callback_query(F.data.startswith("adm:confirm_product:"))
async def confirm_product(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    product_id = int(callback.data.rsplit(":", 1)[1])
    await callback.message.edit_text(
        "⚠️ <b>Удалить товар?</b>\n\nБудут удалены все его разделы и материалы. Это действие нельзя отменить.",
        reply_markup=confirm_keyboard(f"adm:delete_product:{product_id}", f"adm:p:{product_id}"),
    ); await callback.answer()


@router.callback_query(F.data.startswith("adm:delete_product:"))
async def delete_product(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    materials_db.delete_product(int(callback.data.rsplit(":", 1)[1]))
    await callback.message.edit_text("✅ Товар и все связанные материалы удалены.", reply_markup=products_keyboard(admin=True)); await callback.answer()


@router.callback_query(F.data.startswith("adm:confirm_section:"))
async def confirm_section(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1]); section = materials_db.get_section(section_id)
    await callback.message.edit_text(
        "⚠️ <b>Удалить раздел?</b>\n\nВсе материалы внутри него также будут удалены. Это действие нельзя отменить.",
        reply_markup=confirm_keyboard(f"adm:delete_section:{section_id}", f"adm:s:{section_id}"),
    ); await callback.answer()


@router.callback_query(F.data.startswith("adm:delete_section:"))
async def delete_section(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    section_id = int(callback.data.rsplit(":", 1)[1]); section = materials_db.get_section(section_id)
    product_id = section.product_id if section else 0; materials_db.delete_section(section_id)
    await callback.message.edit_text("✅ Раздел удалён.", reply_markup=admin_product_keyboard(product_id)); await callback.answer()


@router.callback_query(F.data.startswith("adm:confirm_material:"))
async def confirm_material(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    material_id = int(callback.data.rsplit(":", 1)[1]); material = materials_db.get_material(material_id)
    await callback.message.edit_text(
        "⚠️ <b>Удалить материал?</b>\n\nЭто действие нельзя отменить.",
        reply_markup=confirm_keyboard(f"adm:delete_material:{material_id}", f"adm:s:{material.section_id}"),
    ); await callback.answer()


@router.callback_query(F.data.startswith("adm:delete_material:"))
async def delete_material(callback: CallbackQuery) -> None:
    if not await require_admin(callback): return
    material_id = int(callback.data.rsplit(":", 1)[1]); material = materials_db.get_material(material_id)
    section_id = material.section_id if material else 0; materials_db.delete_material(material_id)
    await callback.message.edit_text("✅ Материал удалён.", reply_markup=admin_section_keyboard(section_id)); await callback.answer()


@router.message(StateFilter(None))
async def outside_mode(message: Message) -> None:
    await message.answer(
        "Чтобы продолжить, выберите нужный раздел:",
        reply_markup=main_menu(message.from_user.id if message.from_user else None),
    )


async def main() -> None:
    global catalog, materials_db, prices_db, admin_ids
    global active_excel_path, managed_excel_path, price_storage_path
    load_dotenv(BASE_DIR / ".env")
    token = os.getenv("BOT_TOKEN", "").strip()
    if not token:
        raise RuntimeError("Добавьте BOT_TOKEN в файл .env.")
    excel_path = Path(os.getenv("EXCEL_PATH", "managers.xlsx"))
    managed_excel_path = Path(os.getenv("MANAGED_EXCEL_PATH", "data/managers.xlsx"))
    db_path = Path(os.getenv("MATERIALS_DB", "data/materials.sqlite3"))
    prices_db_path = Path(os.getenv("PRICES_DB", "data/prices.sqlite3"))
    price_storage_path = Path(os.getenv("PRICE_STORAGE", "data/prices"))
    if not excel_path.is_absolute(): excel_path = BASE_DIR / excel_path
    if not managed_excel_path.is_absolute(): managed_excel_path = BASE_DIR / managed_excel_path
    if not db_path.is_absolute(): db_path = BASE_DIR / db_path
    if not prices_db_path.is_absolute(): prices_db_path = BASE_DIR / prices_db_path
    if not price_storage_path.is_absolute(): price_storage_path = BASE_DIR / price_storage_path
    admin_ids = {int(value.strip()) for value in os.getenv("ADMIN_IDS", "5533726476").split(",") if value.strip()}
    active_excel_path = managed_excel_path if managed_excel_path.exists() else excel_path
    catalog = Catalog(active_excel_path)
    materials_db = MaterialsDB(db_path)
    for registered_chat in materials_db.list_client_chats():
        cleaned_title = clean_client_title(registered_chat.title)
        if cleaned_title != registered_chat.title:
            materials_db.upsert_client_chat(
                registered_chat.chat_id, cleaned_title, registered_chat.chat_type, registered_chat.is_active
            )
    prices_db = PricesDB(prices_db_path)
    bot = Bot(token=token, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dispatcher = Dispatcher()
    router.message.outer_middleware(AccessMiddleware())
    router.callback_query.outer_middleware(AccessMiddleware())
    dispatcher.include_router(router)
    await bot.delete_webhook(drop_pending_updates=False)
    await dispatcher.start_polling(bot)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
    asyncio.run(main())
